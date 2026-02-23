#!/usr/bin/env python3
"""
PPLX GUI - Graphical User Interface for PPLX File Editor

A user-friendly GUI application for editing pplx XML files, specifically
designed for managing Aux Data fields in pole line engineering files.

Features:
- File selection and browsing
- Visual editing of Aux Data fields
- Automatic output to 'Processed PPLX' folder
- Path memory using JSON configuration
- Batch processing capabilities

Author: AI Assistant
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import json
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import List, Dict
from concurrent.futures import ThreadPoolExecutor
import threading
from datetime import datetime
import csv

# Import our existing PPLX handler
from pplx_handler import PPLXHandler, PPLXBatchProcessor

# Import shared configuration and logic
from pplx_config import (
    analyze_mr_note_for_aux_data,
    extract_scid_from_filename,
    clean_scid_keywords,
    normalize_scid_for_excel_lookup
)

# Excel support
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Constants
POLE_TAG_BLANK = "NO TAG"
PLACEHOLDER_AUX1 = "(Will auto-fill from Excel)"
PLACEHOLDER_AUX2 = "(Will fill from sheet)"

# Aux fields with auto-fill from Excel: (index, config_key, checkbox_label, placeholder, default_checked)
# Checked = fill from sheet (readonly); Unchecked = manual (editable)
AUX_AUTO_FILL_CONFIG = [
    (0, "auto_fill_aux1", "Auto Fill", PLACEHOLDER_AUX1, False),
    (1, "auto_fill_aux2", "Auto Fill", PLACEHOLDER_AUX2, False),
]


def _safe_filename_part(s, extra_chars: str = "") -> str:
    """Strip to alphanumeric and -_ (and optional extra chars) for filename."""
    allowed = "-_" + extra_chars
    return "".join(c for c in str(s) if c.isalnum() or c in allowed)


def _parse_keywords(csv_string: str) -> List[str]:
    """Parse comma-separated string into stripped non-empty keywords."""
    return [kw.strip() for kw in (csv_string or "").split(",") if kw.strip()]


def _set_aux_data_with_log(handler, aux_num: int, value: str, logs: list, prefix: str = "Set") -> bool:
    """Set aux data and append log. Returns success."""
    success = handler.set_aux_data(aux_num, value)
    action = "ERROR: Failed to set" if not success else prefix
    logs.append(f"  {action} Aux Data {aux_num}: {value}")
    return success


def _get_auto_fill_enabled(aux_frame, field_index: int, default: bool = False) -> bool:
    """Return True if aux field is set to auto-fill from Excel."""
    try:
        var = getattr(aux_frame, f"auto_fill_aux{field_index + 1}_var", None)
        return bool(var.get()) if var else default
    except Exception:
        return default


class PPLXConfigManager:
    """Manages application configuration and path memory."""
    
    def __init__(self, config_file: str = "config.json"):
        # Try to find the config file in multiple locations
        self.config_file = self._find_config_file(config_file)
        self.config = self.load_config()
    
    def _find_config_file(self, config_file: str) -> str:
        """Find the config file in multiple possible locations."""
        import sys
        
        # Check if we're running from a PyInstaller bundle
        if getattr(sys, 'frozen', False):
            # Running from executable - look in executable directory
            executable_dir = os.path.dirname(sys.executable)
            possible_paths = [
                os.path.join(executable_dir, config_file),  # Same directory as executable
                os.path.join(os.getcwd(), config_file),  # Current working directory
                config_file,  # Current directory
            ]
            
            # Debug: Print paths being checked (only in development)
            if os.getenv('PPLX_DEBUG'):
                print(f"[DEBUG] Running from executable: {sys.executable}")
                print(f"[DEBUG] Executable directory: {executable_dir}")
                print(f"[DEBUG] Checking config paths:")
                for path in possible_paths:
                    print(f"[DEBUG]   - {path} (exists: {os.path.exists(path)})")
        else:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            possible_paths = [
                os.path.join(script_dir, config_file),
                os.path.join(os.getcwd(), config_file),
                config_file,
            ]
        
        for path in possible_paths:
            if os.path.exists(path):
                return path
        
        fallback_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        return os.path.join(fallback_dir, config_file)
    
    def load_config(self) -> Dict:
        """Load configuration from JSON file."""
        default_config = {
            "last_input_directory": "",
            "last_output_directory": "",
            "window_geometry": "1000x700",
            "recent_files": [],
            "default_aux_values": [""] * 8,
            "configurations": [
                {"name": "Default", "power_label": "POWER"},
                {"name": "OPPD", "power_label": "OPPD"}
            ],
            "selected_config": "OPPD"
        }
        
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                # Merge with defaults to handle new keys
                for key, value in default_config.items():
                    if key not in config:
                        config[key] = value
                # Ensure Default and OPPD configs exist
                configs = config.get("configurations", [])
                if isinstance(configs, list):
                    names = [c.get("name") for c in configs if isinstance(c, dict)]
                    for req in [{"name": "Default", "power_label": "POWER"}, {"name": "OPPD", "power_label": "OPPD"}]:
                        if req["name"] not in names:
                            configs = list(configs) + [req]
                            config["configurations"] = configs
                return config
        except Exception as e:
            print(f"Error loading config: {e}")
        
        return default_config
    
    def save_config(self):
        """Save configuration to JSON file."""
        try:
            with open(self.config_file, 'w') as f:
                json.dump(self.config, f, indent=2)
        except Exception as e:
            print(f"Error saving config: {e}")
    
    def get(self, key: str, default=None):
        """Get configuration value."""
        return self.config.get(key, default)
    
    def set(self, key: str, value):
        """Set configuration value."""
        self.config[key] = value
        self.save_config()
    
    def add_recent_file(self, file_path: str):
        """Add file to recent files list."""
        recent = self.config.get("recent_files", [])
        if file_path in recent:
            recent.remove(file_path)
        recent.insert(0, file_path)
        recent = recent[:10]  # Keep only last 10
        self.set("recent_files", recent)


class PPLXFileListFrame(ttk.Frame):
    """Frame for displaying and managing selected PPLX files."""
    
    def __init__(self, parent, config_manager: PPLXConfigManager, category: str):
        super().__init__(parent)
        self.config_manager = config_manager
        self.category = category
        self.files: List[str] = []
        self.current_folder = ""
        self.source_path = ""
        self.source_type = "folder"
        self.display_name = ""
        self.temp_directory = None
        normalized_category = self.category.lower().replace(" ", "_")
        self.config_key = f"last_{normalized_category}_folder_path"
        
        self.setup_ui()
        
        # Auto-load last used source for this category
        last_source = self.config_manager.get(self.config_key, "")
        if last_source:
            try:
                if last_source.lower().endswith(".zip") and os.path.exists(last_source):
                    self.load_zip_source(last_source, remember=False)
                elif os.path.isdir(last_source):
                    self.load_directory_source(last_source, remember=False)
            except Exception as exc:
                print(f"Warning: Could not auto-load previous source '{last_source}': {exc}")
    
    def setup_ui(self):
        """Setup the file list UI."""
        # Title
        title_label = ttk.Label(
            self,
            text=f"{self.category} PPLX Files",
            font=("Arial", 12, "bold")
        )
        title_label.pack(pady=(0, 10))
        
        # Select folder button
        ttk.Button(
            self,
            text=f"Select {self.category} Folder",
            command=self.select_folder
        ).pack(pady=(0, 10))
        
        # File list with scrollbar
        list_frame = ttk.Frame(self)
        list_frame.pack(fill="both", expand=True)
        
        self.file_listbox = tk.Listbox(list_frame, selectmode=tk.EXTENDED)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.file_listbox.yview)
        self.file_listbox.configure(yscrollcommand=scrollbar.set)
        
        self.file_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Current folder label
        self.folder_label = ttk.Label(
            self,
            text=f"No {self.category} source selected",
            foreground="gray"
        )
        self.folder_label.pack(pady=(5, 0))
        
        # Status label
        self.status_label = ttk.Label(
            self,
            text=f"Select a folder or ZIP containing {self.category} PPLX files"
        )
        self.status_label.pack(pady=(10, 0))
    
    def select_folder(self):
        """Select folder containing PPLX files."""
        initial_dir = self.config_manager.get(self.config_key, "")
        if not initial_dir or not os.path.exists(initial_dir):
            initial_dir = os.getcwd()
        
        zip_source = filedialog.askopenfilename(
            title=f"Select {self.category} ZIP Archive (Cancel to pick a folder)",
            initialdir=initial_dir,
            filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
        )
        
        if zip_source:
            source_path = Path(zip_source)
            if source_path.is_file() and source_path.suffix.lower() == ".zip":
                self.load_zip_source(str(source_path))
            elif source_path.is_dir():
                self.load_directory_source(str(source_path))
            else:
                messagebox.showerror(
                    "Invalid Selection",
                    "Please select a folder or a .zip archive containing PPLX files."
                )
            return
        
        folder = filedialog.askdirectory(
            title=f"Select {self.category} Folder with PPLX Files",
            initialdir=initial_dir,
            mustexist=True
        )
        if folder:
            self.load_directory_source(folder)
    
    def load_folder_files(self):
        """Load all PPLX files from the current folder."""
        if not hasattr(self, 'current_folder') or not self.current_folder:
            return
            
        if not os.path.exists(self.current_folder):
            messagebox.showerror("Error", f"Source path not found: {self.current_folder}")
            return
            
        self.files.clear()
        
        try:
            # Find all pplx files recursively
            pplx_files = []
            for root, _, filenames in os.walk(self.current_folder):
                for filename in filenames:
                    if filename.lower().endswith('.pplx'):
                        file_path = os.path.join(root, filename)
                        pplx_files.append(file_path)
            
            self.files = sorted(pplx_files)
            
            # Update folder label
            source_label = self.display_name or os.path.basename(self.current_folder)
            if self.source_type == "zip":
                source_label = f"{source_label} [ZIP]"
            self.folder_label.config(text=f"Source: {source_label}", foreground="black")
            
            self.update_display()
            
            # Log file count instead of popup
            if self.files:
                print(f"Found {len(self.files)} PPLX files in folder")
            else:
                print("No PPLX files found in selected folder")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error reading folder: {str(e)}")
    
    def clear_files(self):
        """Clear all files from the list."""
        self.cleanup_temp_dir()
        self.files.clear()
        if hasattr(self, 'current_folder'):
            self.current_folder = ""
        self.source_path = ""
        self.display_name = ""
        self.source_type = "folder"
        self.folder_label.config(
            text=f"No {self.category} source selected",
            foreground="gray"
        )
        self.update_display()
    
    def update_display(self):
        """Update the file list display."""
        self.file_listbox.delete(0, tk.END)
        
        for file_path in self.files:
            try:
                display_name = os.path.relpath(file_path, self.current_folder)
            except ValueError:
                display_name = os.path.basename(file_path)
            self.file_listbox.insert(tk.END, display_name.replace("\\", "/"))
        
        count = len(self.files)
        if count > 0:
            self.status_label.config(
                text=f"{count} {self.category} PPLX file{'s' if count != 1 else ''} found"
            )
        else:
            self.status_label.config(
                text=f"Select a folder or ZIP containing {self.category} PPLX files"
            )
    
    def get_files(self) -> List[str]:
        """Get the list of selected files."""
        return self.files.copy()
    
    def get_working_directory(self) -> str:
        """Return the working directory (real folder on disk)."""
        return self.current_folder
    
    def get_source_path(self) -> str:
        """Return the path originally selected by the user (folder or zip)."""
        return self.source_path or self.current_folder
    
    def get_current_folder(self) -> str:
        """Backward-compatible alias for working directory."""
        return self.current_folder
    
    def cleanup_temp_dir(self):
        """Remove any temporary extraction directory."""
        if self.temp_directory and os.path.exists(self.temp_directory):
            shutil.rmtree(self.temp_directory, ignore_errors=True)
        self.temp_directory = None
    
    def destroy(self):
        """Ensure temp data is cleaned up when frame is destroyed."""
        self.cleanup_temp_dir()
        super().destroy()
    
    def load_directory_source(self, folder_path: str, remember: bool = True):
        """Configure frame to use a regular folder."""
        if not os.path.isdir(folder_path):
            messagebox.showerror("Error", f"Folder not found: {folder_path}")
            return
        
        self.cleanup_temp_dir()
        self.source_type = "folder"
        self.source_path = folder_path
        self.display_name = os.path.basename(os.path.normpath(folder_path)) or folder_path
        self.current_folder = folder_path
        
        if remember:
            self.config_manager.set(self.config_key, folder_path)
        
        self.load_folder_files()
    
    def load_zip_source(self, zip_path: str, remember: bool = True):
        """Configure frame to use a ZIP archive, extracting it to a temp directory."""
        if not os.path.exists(zip_path):
            messagebox.showerror("Error", f"ZIP file not found: {zip_path}")
            return
        
        self.cleanup_temp_dir()
        temp_dir = ""
        try:
            temp_dir = tempfile.mkdtemp(prefix=f"pplx_{self.category.lower()}_")
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
        except Exception as exc:
            if temp_dir:
                shutil.rmtree(temp_dir, ignore_errors=True)
            messagebox.showerror("Error", f"Failed to extract ZIP:\n{exc}")
            return
        
        self.temp_directory = temp_dir
        self.source_type = "zip"
        self.source_path = zip_path
        self.display_name = os.path.basename(zip_path)
        self.current_folder = temp_dir
        
        if remember:
            self.config_manager.set(self.config_key, zip_path)
        
        self.load_folder_files()


class AuxDataEditFrame(ttk.Frame):
    """Frame for editing Aux Data fields."""
    
    def __init__(self, parent, config_manager: PPLXConfigManager):
        super().__init__(parent)
        self.config_manager = config_manager
        self.aux_entries = []
        self.ignore_scid_keywords = self.config_manager.get("ignore_scid_keywords", "")
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup the Aux Data editing UI."""
        # Configuration dropdown (top)
        config_frame = ttk.Frame(self)
        config_frame.pack(fill="x", padx=10, pady=(0, 8))
        ttk.Label(config_frame, text="Configuration:").pack(side="left", padx=(0, 6))
        self.config_var = tk.StringVar()
        self.config_combo = ttk.Combobox(config_frame, textvariable=self.config_var, state="readonly", width=25)
        self._refresh_config_combo()
        saved = self.config_manager.get("selected_config", "OPPD")
        names = list(self.config_combo["values"])
        self.config_var.set(saved if saved in names else (names[0] if names else "OPPD"))
        self.config_combo.pack(side="left", padx=(0, 4))
        self.config_combo.bind("<<ComboboxSelected>>", self._on_config_selected)
        ttk.Button(config_frame, text="Add Config...", command=self._add_config_dialog).pack(side="left", padx=(4, 0))
        
        # Excel file selection (top of panel)
        excel_frame = ttk.Frame(self)
        excel_frame.pack(fill="x", padx=10, pady=(0, 12))
        
        ttk.Label(excel_frame, text="Node-Section-Connection File:").pack(side="left", padx=(0, 4))
        self.excel_file_var = tk.StringVar()
        self.excel_file_var.set(self.config_manager.get("excel_file_path", "No file selected"))
        
        excel_label = ttk.Label(excel_frame, textvariable=self.excel_file_var, foreground="blue")
        excel_label.pack(side="left", fill="x", expand=True, padx=(5, 5))
        
        ttk.Button(excel_frame, text="Browse", command=self.select_excel_file).pack(side="left")
        
        # Title
        title_label = ttk.Label(self, text="Aux Data Fields", font=("Arial", 12, "bold"))
        title_label.pack(pady=(0, 10))
        
        # Keyword configuration frame
        owners_frame = ttk.LabelFrame(self, text="Keyword Configuration", padding=10)
        owners_frame.pack(fill="x", padx=10, pady=(10, 0))
        owners_frame.columnconfigure(1, weight=1)
        
        self.comm_owners_var = tk.StringVar(value=self.config_manager.get("comm_owners", ""))
        self.power_owners_var = tk.StringVar(value=self.config_manager.get("power_owners", ""))
        self.pco_keywords_var = tk.StringVar(value=self.config_manager.get("pco_keywords", ""))
        self.aux5_keywords_var = tk.StringVar(value=self.config_manager.get("aux5_keywords", ""))

        keyword_fields = [
            ("Comm Keywords:", self.comm_owners_var),
            ("Power Keywords:", self.power_owners_var),
            ("PCO Keywords:", self.pco_keywords_var),
            ("Riser Keywords:", self.aux5_keywords_var)
        ]

        for row, (label_text, var) in enumerate(keyword_fields):
            ttk.Label(owners_frame, text=label_text, width=18, anchor="w").grid(row=row, column=0, sticky="w", pady=2, padx=(0, 6))
            entry = ttk.Entry(owners_frame, textvariable=var)
            entry.grid(row=row, column=1, sticky="ew", pady=2)
            entry.bind('<FocusOut>', self.save_owner_config)
        
        # Define the 5 Aux Data fields with their descriptions
        self.aux_fields = [
            ("Aux Data 1", "Pole Owner", True),     # User editable or auto-filled
            ("Aux Data 2", "Pole Tag", True),       # User editable or auto-filled
            ("Aux Data 3", "Condition", True),      # User editable  
            ("Aux Data 4", "Make Ready Type", False), # Auto-filled
            ("Aux Data 5", "Proposed Riser", False)  # Auto-filled
        ]
        
        # Create entry fields for editable fields and labels for auto-filled
        fields_frame = ttk.Frame(self)
        fields_frame.pack(fill="x", padx=10)
        
        saved_values = self.config_manager.get("last_aux_values", {})
        
        for i, (field_name, description, is_editable) in enumerate(self.aux_fields):
            row_frame = ttk.Frame(fields_frame)
            row_frame.pack(fill="x", pady=2)
            
            # Label with description in parentheses
            label_text = f"{field_name} ({description}):"
            label = ttk.Label(row_frame, text=label_text, width=30)
            label.pack(side="left")
            
            if is_editable:
                # Create entry for user input
                entry = ttk.Entry(row_frame)
                entry.pack(side="left", fill="x", expand=True, padx=(5, 0))
                
                # Load saved value if available first
                saved_key = f"aux_data_{i+1}"
                if i == 2:
                    entry.insert(0, "Auto (EXISTING/PROPOSED)")
                    entry.config(state="readonly")
                elif saved_key in saved_values:
                    entry.insert(0, saved_values[saved_key])
                
                # Add auto-fill checkbox for configured aux fields
                for idx, config_key, label, placeholder, default_checked in AUX_AUTO_FILL_CONFIG:
                    if i == idx:
                        var = tk.BooleanVar()
                        setattr(self, f"auto_fill_aux{idx + 1}_var", var)
                        ttk.Checkbutton(
                            row_frame, text=label, variable=var,
                            command=lambda i=idx: self._toggle_aux_auto_fill(i)
                        ).pack(side="left", padx=(5, 0))
                        checkbox_state = self.config_manager.get(config_key, default_checked)
                        var.set(checkbox_state)
                        # Checked = fill from sheet (readonly placeholder); Unchecked = manual (editable)
                        show_placeholder = checkbox_state
                        if show_placeholder:
                            entry.config(state="readonly")
                            entry.delete(0, tk.END)
                            entry.insert(0, placeholder)
                        break
                
                # Bind to auto-save when value changes
                if i != 2:
                    entry.bind('<FocusOut>', self.auto_save_values)
                    entry.bind('<KeyRelease>', self.auto_save_values)
                
                self.aux_entries.append(entry)
            else:
                # Show "(Auto Filled)" for non-editable fields
                auto_label = ttk.Label(row_frame, text="(Auto Filled)", foreground="gray", style="Gray.TLabel")
                auto_label.pack(side="left", fill="x", expand=True, padx=(5, 0))
                self.aux_entries.append(auto_label)
    
    def _refresh_config_combo(self):
        """Refresh the configuration dropdown from config."""
        configs = self.config_manager.get("configurations", [])
        names = [c["name"] for c in configs if isinstance(c, dict)]
        self.config_combo["values"] = names
    
    def _on_config_selected(self, event=None):
        """Save selected config when user changes dropdown."""
        val = self.config_var.get()
        if val:
            self.config_manager.set("selected_config", val)
    
    def _add_config_dialog(self):
        """Open dialog to add a new configuration."""
        dialog = tk.Toplevel(self)
        dialog.title("Add Configuration")
        dialog.geometry("320x140")
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()
        
        ttk.Label(dialog, text="Config Name:").pack(anchor="w", padx=15, pady=(15, 2))
        name_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=name_var, width=35).pack(fill="x", padx=15, pady=(0, 10))
        
        ttk.Label(dialog, text="Power Label (replaces 'POWER' in Aux Data 4):").pack(anchor="w", padx=15, pady=(5, 2))
        power_var = tk.StringVar(value="POWER")
        ttk.Entry(dialog, textvariable=power_var, width=35).pack(fill="x", padx=15, pady=(0, 15))
        
        def do_add():
            name = name_var.get().strip()
            power = power_var.get().strip() or "POWER"
            if not name:
                messagebox.showwarning("Invalid", "Please enter a configuration name.", parent=dialog)
                return
            configs = self.config_manager.get("configurations", [])
            if any(c.get("name") == name for c in configs if isinstance(c, dict)):
                messagebox.showwarning("Duplicate", f"Configuration '{name}' already exists.", parent=dialog)
                return
            configs.append({"name": name, "power_label": power})
            self.config_manager.set("configurations", configs)
            self._refresh_config_combo()
            self.config_var.set(name)
            self.config_manager.set("selected_config", name)
            dialog.destroy()
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))
        ttk.Button(btn_frame, text="Add", command=do_add).pack(side="right", padx=(5, 0))
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side="right")
    
    def get_selected_config_power_label(self) -> str:
        """Return the power_label for the selected config (used for Aux Data 4 substitution)."""
        selected = self.config_var.get()
        configs = self.config_manager.get("configurations", [])
        for c in configs:
            if isinstance(c, dict) and c.get("name") == selected:
                return c.get("power_label", "POWER")
        return "POWER"
    
    def select_excel_file(self):
        """Select Excel file for data lookup."""
        from tkinter import filedialog
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_file_var.set(file_path)
            self.config_manager.set("excel_file_path", file_path)
    
    def get_excel_path(self) -> str:
        """Return the currently selected Excel file path."""
        value = self.excel_file_var.get()
        if value and value != "No file selected":
            return value
        return self.config_manager.get("excel_file_path", "")
    
    def auto_save_values(self, event=None):
        """Auto-save current values when they change (skips placeholders and category-driven fields)."""
        values = {}
        entry_index = 0
        placeholders = (PLACEHOLDER_AUX1, PLACEHOLDER_AUX2, "Auto (EXISTING/PROPOSED)")
        for i, (_, _, is_editable) in enumerate(self.aux_fields):
            if not is_editable:
                entry_index += 1
                continue
            entry = self.aux_entries[entry_index]
            entry_index += 1
            if hasattr(entry, 'get'):
                value = entry.get().strip()
                if value and value not in placeholders:
                    values[f"aux_data_{i + 1}"] = value
        self.config_manager.set("last_aux_values", values)
    
    def _toggle_aux_auto_fill(self, field_index: int):
        """Toggle auto-fill from Excel. Checked = fill (readonly); Unchecked = manual (editable)."""
        config_key, _, placeholder, *_ = next(c for c in AUX_AUTO_FILL_CONFIG if c[0] == field_index)
        var = getattr(self, f"auto_fill_aux{field_index + 1}_var")
        state = var.get()
        self.config_manager.set(config_key, state)
        entry = self.aux_entries[field_index]
        show_readonly_placeholder = state
        if show_readonly_placeholder:
            entry.config(state="normal")
            entry.delete(0, tk.END)
            entry.insert(0, placeholder)
            entry.config(state="readonly")
        else:
            entry.config(state="normal")

    def save_owner_config(self, event=None):
        """Save owner configuration when values change."""
        self.config_manager.set("comm_owners", self.comm_owners_var.get())
        self.config_manager.set("power_owners", self.power_owners_var.get())
        if hasattr(self, 'pco_keywords_var'):
            self.config_manager.set("pco_keywords", self.pco_keywords_var.get())
        if hasattr(self, 'aux5_keywords_var'):
            self.config_manager.set("aux5_keywords", self.aux5_keywords_var.get())
    
    def analyze_mr_note(self, mr_note: str) -> tuple:
        """Analyze mr_note to determine Aux Data 4 and 5 values."""
        def get_keywords(var_attr: str) -> List[str]:
            var = getattr(self, var_attr, None)
            return _parse_keywords(var.get() if var else "")

        return analyze_mr_note_for_aux_data(
            mr_note,
            comm_keywords=get_keywords("comm_owners_var"),
            power_keywords=get_keywords("power_owners_var"),
            pco_keywords=get_keywords("pco_keywords_var"),
            aux5_keywords=get_keywords("aux5_keywords_var")
        )
    
    def get_aux_values(self) -> Dict[int, str]:
        """Get the current Aux Data values for processing (excludes auto-fill placeholders)."""
        values = {}
        placeholders = (PLACEHOLDER_AUX1, PLACEHOLDER_AUX2)
        entry_index = 0
        for i, (_, _, is_editable) in enumerate(self.aux_fields):
            if not is_editable:
                entry_index += 1
                continue
            # Skip when filling from Excel (checkbox checked)
            if i in {c[0] for c in AUX_AUTO_FILL_CONFIG} and _get_auto_fill_enabled(self, i):
                entry_index += 1
                continue
            entry = self.aux_entries[entry_index]
            entry_index += 1
            if hasattr(entry, 'get'):
                value = entry.get().strip()
                if value and value not in placeholders:
                    values[i + 1] = value
        return values
    
    def set_readonly_field(self, field_index: int, value: str):
        """Set value in a readonly field (0-based index)."""
        if 0 <= field_index < len(self.aux_entries):
            entry = self.aux_entries[field_index]
            # Temporarily enable readonly fields to set value
            was_readonly = str(entry.cget('state')) == 'readonly'
            if was_readonly:
                entry.config(state='normal')
            entry.delete(0, tk.END)
            entry.insert(0, value)
            if was_readonly:
                entry.config(state='readonly')
    
    
    def load_excel_data(self, log_callback=None) -> Dict[str, Dict]:
        """Load and filter Excel data based on node_type=pole and pole_status!=underground"""
        excel_path = self.config_manager.get("excel_file_path", "")
        if log_callback:
            log_callback(f"Looking for Excel file at: {excel_path}")
            log_callback(f"Current working directory: {os.getcwd()}")
        
        if not excel_path or not os.path.exists(excel_path):
            if log_callback:
                log_callback(f"Excel Error: File not found or path empty: {excel_path}")
                log_callback(f"Please check the Excel file path in the configuration")
            return {}
        
        if not EXCEL_AVAILABLE:
            if log_callback:
                log_callback("Excel Support: openpyxl library not available. Please install it.")
            return {}
        
        try:
            workbook = openpyxl.load_workbook(excel_path)
            if 'nodes' not in workbook.sheetnames:
                if log_callback:
                    log_callback("Excel Error: No 'nodes' sheet found in Excel file")
                return {}
            
            sheet = workbook['nodes']
            data = {}
            
            # Get headers
            headers = {}
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header:
                    headers[header.lower()] = col
            
            # Check required columns
            required_cols = ['scid', 'node_type', 'pole_status']
            optional_cols = ['mr_note', 'pole_tag_company', 'pole_tag_tagtext']

            missing_cols = [col for col in required_cols if col not in headers]
            if missing_cols:
                if log_callback:
                    log_callback(f"Excel Error: Missing required columns in Excel file: {missing_cols}")
                return {}

            # Check optional columns and log warnings if missing
            missing_optional = [col for col in optional_cols if col not in headers]
            if missing_optional:
                if log_callback:
                    log_callback(f"Excel Warning: Optional columns not found (will use defaults): {missing_optional}")
            
            # Read data with filtering
            valid_count = 0
            total_count = 0
            skipped_no_scid = 0
            skipped_filters = 0

            for row in range(2, sheet.max_row + 1):
                total_count += 1
                scid = sheet.cell(row=row, column=headers['scid']).value
                node_type = sheet.cell(row=row, column=headers['node_type']).value
                pole_status = sheet.cell(row=row, column=headers['pole_status']).value

                # Skip rows without SCID
                if not scid:
                    skipped_no_scid += 1
                    continue

                # Apply filters: node_type = pole and pole_status != 'underground' (allow missing pole_status)
                if (node_type and str(node_type).lower() == 'pole' and
                    (pole_status is None or str(pole_status).lower() != 'underground')):

                    # Store all row data for this SCID
                    row_data = {}
                    for col_name, col_num in headers.items():
                        cell_value = sheet.cell(row=row, column=col_num).value
                        row_data[col_name] = str(cell_value) if cell_value is not None else ""

                    data[str(scid)] = row_data
                    valid_count += 1
                else:
                    skipped_filters += 1
            
            workbook.close()
            if log_callback:
                log_callback(f"Excel loaded successfully: {valid_count} valid pole entries")
                log_callback(f"  Total rows: {total_count} | Skipped (no SCID): {skipped_no_scid} | Skipped (filters): {skipped_filters}")
            return data
            
        except Exception as e:
            if log_callback:
                log_callback(f"Excel Error: Error reading Excel file: {str(e)}")
            return {}
    
    def get_valid_scids(self) -> set:
        """Get set of valid SCIDs from Excel data"""
        excel_data = self.load_excel_data()
        return set(excel_data.keys())
    
    def apply_auto_fill_logic(self, scid: str, excel_data: Dict[str, Dict] = None, log_callback=None):
        """Apply logic to auto-fill Aux Data 2, 4, and 5 based on SCID and Excel data."""
        if excel_data is None:
            excel_data = self.load_excel_data(log_callback=log_callback)
        
        if scid not in excel_data:
            # Set default values if SCID not found
            self.set_readonly_field(1, "NO DATA")  # Aux Data 2
            self.set_readonly_field(3, "NO DATA")  # Aux Data 4  
            self.set_readonly_field(4, "NO DATA")  # Aux Data 5
            return
        
        row_data = excel_data[scid]
        
        # Auto-fill logic - you can customize these based on Excel columns
        # Aux Data 2 (Pole Tag) - using SCID as tag for now
        pole_tag = f"POLE_{scid}"
        self.set_readonly_field(1, pole_tag)
        
        # Aux Data 4 (Make Ready Type) - placeholder logic
        # You can map this to specific Excel columns
        make_ready_type = row_data.get('make_ready_type', 'STANDARD')
        self.set_readonly_field(3, make_ready_type)
        
        # Aux Data 5 (Proposed) - placeholder logic  
        # You can map this to specific Excel columns
        proposed = row_data.get('proposed_status', 'NEW')
        self.set_readonly_field(4, proposed)



class ProcessingFrame(ttk.Frame):
    """Frame for processing controls and output."""
    
    def __init__(
        self,
        parent,
        config_manager: PPLXConfigManager,
        existing_frame: PPLXFileListFrame,
        proposed_frame: PPLXFileListFrame,
        aux_frame: AuxDataEditFrame
    ):
        super().__init__(parent)
        self.config_manager = config_manager
        self.existing_frame = existing_frame
        self.proposed_frame = proposed_frame
        self.aux_frame = aux_frame
        self.is_processing = False
        self.active_output_root = ""
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup the processing UI."""
        # Title
        title_label = ttk.Label(self, text="Processing", font=("Arial", 12, "bold"))
        title_label.pack(pady=(0, 10))
        
        # Output information
        info_frame = ttk.Frame(self)
        info_frame.pack(fill="x", pady=(0, 10))
        
        info_text = ttk.Label(
            info_frame, 
            text="Processed files will be saved under 'Processed PPLX' in your Downloads folder\nwith separate 'EXISTING' and 'PROPOSED' folders",
            foreground="gray",
            justify="center"
        )
        info_text.pack()
        
        # Process button
        self.process_button = ttk.Button(self, text="Process Files", command=self.start_processing)
        self.process_button.pack(pady=(0, 10))
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill="x", pady=(0, 10))
        
        # Output text area
        output_frame = ttk.LabelFrame(self, text="Processing Log", padding=5)
        output_frame.pack(fill="both", expand=True)
        
        self.output_text = scrolledtext.ScrolledText(output_frame, height=10, wrap=tk.WORD)
        self.output_text.pack(fill="both", expand=True)
    
    def log_message(self, message: str):
        """Add a message to the output log."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {message}\n"
        
        self.output_text.insert(tk.END, formatted_message)
        self.output_text.see(tk.END)
        self.output_text.update()
    
    def start_processing(self):
        """Start the file processing in a separate thread."""
        if self.is_processing:
            return
        
        existing_files = self.existing_frame.get_files()
        proposed_files = self.proposed_frame.get_files()
        if not existing_files and not proposed_files:
            self.log_message("ERROR: Please select at least one folder containing PPLX files")
            return
        
        excel_path = ""
        if hasattr(self.aux_frame, "get_excel_path") and callable(getattr(self.aux_frame, "get_excel_path")):
            excel_path = self.aux_frame.get_excel_path()
        if not excel_path or not os.path.exists(excel_path):
            self.log_message("ERROR: Please select a valid Excel file before processing.")
            messagebox.showerror("Excel Required", "Please select a valid Excel file in the Aux Data panel before processing.")
            return
        
        aux_values = self.aux_frame.get_aux_values()
        if not aux_values:
            self.log_message("WARNING: No Aux Data values specified. Files will be copied with auto-filled data only.")
        
        # Disable the process button
        self.is_processing = True
        self.process_button.config(state="disabled", text="Processing...")
        
        # Clear the output log
        self.output_text.delete(1.0, tk.END)
        
        # Prepare category data for processing
        category_data = [
            {
                "name": "EXISTING",
                "files": existing_files,
                "source_folder": self.existing_frame.get_source_path()
            },
            {
                "name": "PROPOSED",
                "files": proposed_files,
                "source_folder": self.proposed_frame.get_source_path()
            }
        ]
        
        # Determine output root on Desktop
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        downloads_dir = Path.home() / "Downloads"
        processed_base = downloads_dir / "Processed PPLX"
        processed_base.mkdir(parents=True, exist_ok=True)
        prefix = "O-Calcs"
        if excel_path:
            base_name = os.path.basename(excel_path)
            prefix_candidate = base_name.split(" ")[0].strip()
            if not prefix_candidate:
                prefix_candidate = Path(base_name).stem
            if prefix_candidate:
                prefix = f"{prefix_candidate}_O-Calcs"
        
        output_root = processed_base / f"{prefix}_{timestamp}"
        self.active_output_root = str(output_root)
        
        # Start processing in a separate thread
        thread = threading.Thread(
            target=self.process_files,
            args=(category_data, aux_values, self.active_output_root)
        )
        thread.daemon = True
        thread.start()
    
    def process_files(self, category_data: List[Dict], aux_values: Dict[int, str], output_root: str):
        """Process the selected files for each category with SCID filtering."""
        try:
            os.makedirs(output_root, exist_ok=True)
            
            # Load Excel data and get valid SCIDs
            excel_data = self.aux_frame.load_excel_data(log_callback=self.log_message)
            valid_scids = set(excel_data.keys()) if excel_data else set()
            
            total_files = sum(len(category["files"]) for category in category_data)
            if total_files == 0:
                self.log_message("No PPLX files found to process.")
                return
            
            self.log_message(f"Starting processing of {total_files} files across categories")
            self.log_message(f"Output root directory: {output_root}")
            
            if excel_data:
                self.log_message(f"{len(valid_scids)} non-underground poles found in Excel data")
            else:
                self.log_message("No Excel data loaded - processing all files")
            
            if aux_values:
                self.log_message(f"User Aux Data values (shared): {aux_values}")
            
            power_label = self.aux_frame.get_selected_config_power_label()
            if power_label != "POWER":
                self.log_message(f"Aux Data 4: replacing POWER with '{power_label}'")
            
            processed_count = 0
            summary: Dict[str, Dict[str, object]] = {}
            
            for category in category_data:
                name = category["name"]
                files = category["files"]
                source_folder = category.get("source_folder") or ""
                
                output_dir = os.path.join(output_root, name)
                os.makedirs(output_dir, exist_ok=True)
                
                summary[name] = {
                    "successful": 0,
                    "failed": 0,
                    "skipped": 0,
                    "csv_path": "",
                    "output_dir": output_dir
                }
                
                if not files:
                    self.log_message(f"\nCategory: {name} (no files selected, directory created)")
                    continue
                
                condition_value = name
                category_aux_values = {k: v for k, v in aux_values.items() if k != 3}
                csv_data = []
                
                self.log_message(f"\nCategory: {name}")
                if source_folder:
                    self.log_message(f" Source: {source_folder}")
                self.log_message(f" Processing {len(files)} file{'s' if len(files) != 1 else ''}")

                ignore_keywords = getattr(self.aux_frame, "ignore_scid_keywords", "")


                auto_fill_aux1_enabled = _get_auto_fill_enabled(self.aux_frame, 0, False)
                auto_fill_aux2_enabled = _get_auto_fill_enabled(self.aux_frame, 1, False)
                
                def _get_str_var(name):
                    try:
                        var = getattr(self.aux_frame, name, None)
                        return (var.get() if var else "") or ""
                    except Exception:
                        return ""

                keyword_payload = {
                    "comm_keywords": _parse_keywords(_get_str_var("comm_owners_var")),
                    "power_keywords": _parse_keywords(_get_str_var("power_owners_var")),
                    "pco_keywords": _parse_keywords(_get_str_var("pco_keywords_var")),
                    "aux5_keywords": _parse_keywords(_get_str_var("aux5_keywords_var"))
                }
                
                def process_single_file(task):
                    """Process a single file in a worker thread."""
                    index, file_path = task
                    logs = []
                    csv_row = None
                    status = "success"
                    filename = os.path.basename(file_path)
                    
                    try:
                        scid = extract_scid_from_filename(filename)
                        pole_number = extract_scid_from_filename(filename)
                        clean_pole_number = clean_scid_keywords(
                            pole_number,
                            ignore_keywords
                        )
                        
                        if excel_data and scid not in valid_scids:
                            logs.append(f"Skipping {filename}: SCID '{scid}' not found in Excel data")
                            return {
                                "index": index,
                                "status": "skipped",
                                "logs": logs,
                                "csv_row": None
                            }
                        
                        logs.append(
                            f"Processing: {filename} (SCID: {scid}, Pole Number: {pole_number} -> {clean_pole_number})"
                        )
                        
                        handler = PPLXHandler(file_path)
                        
                        if category_aux_values:
                            for aux_num, value in category_aux_values.items():
                                _set_aux_data_with_log(handler, aux_num, value, logs, "Set")

                        _set_aux_data_with_log(handler, 3, condition_value, logs, "Auto-set")
                        
                        pole_tag = POLE_TAG_BLANK
                        mr_note = ""

                        if excel_data and scid in excel_data:
                            row_data = excel_data[scid]

                            if auto_fill_aux1_enabled:
                                pole_owner = row_data.get('pole_tag_company', '')
                                if pole_owner:
                                    _set_aux_data_with_log(handler, 1, pole_owner, logs, "Auto-filled")

                            if auto_fill_aux2_enabled:
                                excel_pole_tag = row_data.get('pole_tag_tagtext', '').strip()
                                pole_tag = excel_pole_tag if excel_pole_tag else POLE_TAG_BLANK
                                _set_aux_data_with_log(handler, 2, pole_tag, logs, "Auto-filled")
                            else:
                                pole_tag = category_aux_values.get(2, POLE_TAG_BLANK)
                                _set_aux_data_with_log(handler, 2, pole_tag, logs, "Set (manual)")

                            mr_note = row_data.get('mr_note', '')
                            aux_data_4, aux_data_5 = analyze_mr_note_for_aux_data(
                                mr_note,
                                comm_keywords=keyword_payload["comm_keywords"],
                                power_keywords=keyword_payload["power_keywords"],
                                pco_keywords=keyword_payload["pco_keywords"],
                                aux5_keywords=keyword_payload["aux5_keywords"]
                            )
                            # Replace POWER with config's power_label (e.g. OPPD) in Aux Data 4
                            if power_label != "POWER":
                                aux_data_4 = aux_data_4.replace("POWER", power_label)

                            _set_aux_data_with_log(handler, 4, aux_data_4, logs, "Auto-filled")
                            if mr_note:
                                logs.append(f"    Based on mr_note: {mr_note[:50]}{'...' if len(mr_note) > 50 else ''}")

                            _set_aux_data_with_log(handler, 5, aux_data_5, logs, "Auto-filled")
                        else:
                            pole_tag = category_aux_values.get(2, POLE_TAG_BLANK) if not auto_fill_aux2_enabled else pole_tag
                            _set_aux_data_with_log(handler, 2, pole_tag, logs, "Set (fallback)")
                        
                        if excel_data and scid in excel_data:
                            if auto_fill_aux2_enabled:
                                pole_tag = excel_data[scid].get('pole_tag_tagtext', pole_tag)
                            mr_note = excel_data[scid].get('mr_note', mr_note)
                        
                        final_aux_data = handler.get_aux_data()
                        
                        aux_data_4 = final_aux_data.get('Aux Data 4', '')
                        if aux_data_4 == 'PCO':
                            clean_pole_number = f"{clean_pole_number} PCO"
                            logs.append(
                                f"  Aux Data 4 is 'PCO', appending to pole number: {clean_pole_number}"
                            )
                        
                        clean_pole_number_safe = _safe_filename_part(
                            clean_pole_number, ". " if aux_data_4 == "PCO" else ""
                        )
                        clean_pole_tag = _safe_filename_part(pole_tag, " ")
                        clean_condition = _safe_filename_part(condition_value)
                        
                        new_filename = f"{clean_pole_number_safe}_{clean_pole_tag}_{clean_condition}.pplx"
                        output_file = os.path.join(output_dir, new_filename)
                        
                        handler.set_pole_attribute('Pole Number', clean_pole_number)
                        logs.append(f"  Set Pole Number: {clean_pole_number}")
                        
                        description_override = os.path.splitext(new_filename)[0]
                        handler.set_pole_attribute('DescriptionOverride', description_override)
                        logs.append(f"  Set DescriptionOverride: {description_override}")
                        
                        handler.save_file(output_file)
                        logs.append(f"  Saved: {os.path.basename(output_file)}")
                        
                        csv_row = {
                            'File Name': filename,
                            'MR Note': mr_note,
                            'Aux Data 1': final_aux_data.get('Aux Data 1', 'Unset'),
                            'Aux Data 2': final_aux_data.get('Aux Data 2', 'Unset'),
                            'Aux Data 3': final_aux_data.get('Aux Data 3', 'Unset'),
                            'Aux Data 4': final_aux_data.get('Aux Data 4', 'Unset'),
                            'Aux Data 5': final_aux_data.get('Aux Data 5', 'Unset')
                        }
                    
                    except Exception as e:
                        logs.append(f"  Error processing {filename}: {str(e)}")
                        status = "failed"
                    
                    return {
                        "index": index,
                        "status": status,
                        "logs": logs,
                        "csv_row": csv_row
                    }
                
                max_workers = min(8, max(1, (os.cpu_count() or 1)))
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    for result in executor.map(process_single_file, enumerate(files)):
                        for entry in result["logs"]:
                            self.log_message(entry)
                        
                        status = result["status"]
                        if status == "success":
                            summary[name]["successful"] += 1
                            if result["csv_row"]:
                                csv_data.append(result["csv_row"])
                        elif status == "failed":
                            summary[name]["failed"] += 1
                        elif status == "skipped":
                            summary[name]["skipped"] += 1
                        
                        processed_count += 1
                        progress = (processed_count / total_files) * 100
                        self.progress_var.set(progress)
                
                if csv_data:
                    csv_file_path = os.path.join(output_dir, "log.csv")
                    try:
                        with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
                            fieldnames = [
                                'File Name',
                                'MR Note',
                                'Aux Data 1',
                                'Aux Data 2',
                                'Aux Data 3',
                                'Aux Data 4',
                                'Aux Data 5'
                            ]
                            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                            writer.writeheader()
                            writer.writerows(csv_data)
                        
                        self.log_message(f"{name} CSV log file saved: {csv_file_path}")
                        summary[name]["csv_path"] = csv_file_path
                    except Exception as e:
                        self.log_message(f"Error saving {name} CSV log file: {str(e)}")
            
            self.progress_var.set(100)
            
            self.log_message(f"\nProcessing complete!")
            for name, stats in summary.items():
                self.log_message(
                    f"{name} - Successful: {stats['successful']}, Failed: {stats['failed']}, Skipped: {stats['skipped']}"
                )
                if stats.get("csv_path"):
                    self.log_message(f"{name} CSV: {stats['csv_path']}")
                self.log_message(f"{name} Output: {stats['output_dir']}")
            
            self.log_message(f"Output root directory: {output_root}")
        
            try:
                if os.name == "nt":
                    os.startfile(output_root)
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", output_root])
                else:
                    subprocess.Popen(["xdg-open", output_root])
            except Exception as open_err:
                self.log_message(f"Info: Unable to open output folder automatically: {open_err}")
        
        except Exception as e:
            self.log_message(f"Critical error: {str(e)}")
            messagebox.showerror("Error", f"Processing failed: {str(e)}")
        
        finally:
            self.is_processing = False
            self.process_button.config(state="normal", text="Process Files")


class PPLXGUIApp:
    """Main PPLX GUI Application."""
    
    def __init__(self):
        self.config_manager = PPLXConfigManager()
        self.root = tk.Tk()
        self.setup_window()
        self.setup_ui()
    
    def setup_window(self):
        """Setup the main window."""
        self.root.title("PPLX File Editor - Pole Line Engineering Tools")
        self.root.geometry(self.config_manager.get("window_geometry", "1000x700"))
        
        # Set minimum size
        self.root.minsize(800, 600)
        
        # Configure style
        style = ttk.Style()
        if "clam" in style.theme_names():
            style.theme_use("clam")
        
        # Bind window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def setup_ui(self):
        """Setup the main UI."""
        # Create main container with padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Create paned window for resizable sections
        paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned.pack(fill="both", expand=True)
        
        # Left panel - File selection and Aux Data editing
        left_panel = ttk.Frame(paned)
        paned.add(left_panel, weight=1)
        
        # Existing and Proposed folder selectors side by side
        file_lists_container = ttk.Frame(left_panel)
        file_lists_container.pack(fill="both", expand=True, pady=(0, 10))
        file_lists_container.columnconfigure(0, weight=1)
        file_lists_container.columnconfigure(1, weight=1)
        file_lists_container.rowconfigure(0, weight=1)
        
        self.existing_frame = PPLXFileListFrame(file_lists_container, self.config_manager, category="EXISTING")
        self.existing_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        
        self.proposed_frame = PPLXFileListFrame(file_lists_container, self.config_manager, category="PROPOSED")
        self.proposed_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        
        # Aux data editing frame
        self.aux_frame = AuxDataEditFrame(left_panel, self.config_manager)
        self.aux_frame.pack(fill="x")
        
        # Right panel - Processing controls and output
        right_panel = ttk.Frame(paned)
        paned.add(right_panel, weight=1)
        
        self.processing_frame = ProcessingFrame(
            right_panel,
            self.config_manager,
            self.existing_frame,
            self.proposed_frame,
            self.aux_frame
        )
        self.processing_frame.pack(fill="both", expand=True)
    
    def show_batch_report(self):
        """Show batch report dialog."""
        files = self.existing_frame.get_files() + self.proposed_frame.get_files()
        if not files:
            messagebox.showwarning("No Files", "Please select PPLX files first")
            return
        
        # Create a new window for the report
        report_window = tk.Toplevel(self.root)
        report_window.title("Batch Report")
        report_window.geometry("600x400")
        
        # Create report text
        report_text = scrolledtext.ScrolledText(report_window, wrap=tk.WORD)
        report_text.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Generate report
        report_text.insert(tk.END, f"PPLX Files Batch Report\n")
        report_text.insert(tk.END, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        report_text.insert(tk.END, f"Total Files: {len(files)}\n\n")
        
        for i, file_path in enumerate(files, 1):
            try:
                handler = PPLXHandler(file_path)
                info = handler.get_file_info()
                aux_data = handler.get_aux_data()
                
                report_text.insert(tk.END, f"{i}. {os.path.basename(file_path)}\n")
                report_text.insert(tk.END, f"   Date: {info.get('date', 'Unknown')}\n")
                report_text.insert(tk.END, f"   User: {info.get('user', 'Unknown')}\n")
                
                # Show non-default aux data
                non_default = {k: v for k, v in aux_data.items() if v != 'Unset'}
                if non_default:
                    report_text.insert(tk.END, f"   Aux Data: {non_default}\n")
                else:
                    report_text.insert(tk.END, f"   Aux Data: All unset\n")
                
                report_text.insert(tk.END, "\n")
                
            except Exception as e:
                report_text.insert(tk.END, f"{i}. {os.path.basename(file_path)}\n")
                report_text.insert(tk.END, f"   Error: {str(e)}\n\n")
    
    def export_structure(self):
        """Export XML structure of selected file."""
        files = self.existing_frame.get_files() + self.proposed_frame.get_files()
        if not files:
            messagebox.showwarning("No Files", "Please select a PPLX file first")
            return
        
        # Use first file for structure export
        file_path = files[0]
        
        # Ask for output location
        output_file = filedialog.asksaveasfilename(
            title="Export Structure to JSON",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if output_file:
            try:
                handler = PPLXHandler(file_path)
                handler.export_structure_to_json(output_file)
                messagebox.showinfo("Success", f"Structure exported to:\n{output_file}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export structure:\n{str(e)}")
    
    def on_closing(self):
        """Handle application closing."""
        # Save window geometry
        self.config_manager.set("window_geometry", self.root.geometry())
        
        # Clean up any temporary data
        if hasattr(self, "existing_frame"):
            self.existing_frame.cleanup_temp_dir()
        if hasattr(self, "proposed_frame"):
            self.proposed_frame.cleanup_temp_dir()
        
        # Close the application
        self.root.destroy()
    
    def run(self):
        """Run the application."""
        self.root.mainloop()


def main():
    """Main function to run the GUI application."""
    try:
        app = PPLXGUIApp()
        app.run()
    except Exception as e:
        print(f"Error starting application: {e}")
        messagebox.showerror("Startup Error", f"Failed to start application:\n{str(e)}")


if __name__ == "__main__":
    main() 