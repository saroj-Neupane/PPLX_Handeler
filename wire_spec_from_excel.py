"""
Wire spec between pole pairs from Excel (nodes + connections) and OPPD shapefiles.
ElectricLine: Primary (d_masterma), Neutral (d_neutralm).
S_ElectricLine: Secondary (d_masterma).
"""
import math
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import shapefile  # pyshp
from pyproj import CRS, Transformer

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ---------- Coordinate transform & distance (same as tmp_query_wire_between_points) ----------
def load_crs_from_prj(prj_path: Path) -> CRS:
    return CRS.from_wkt(prj_path.read_text())


def transformer_wgs84_to_layer(prj_path: Path) -> Transformer:
    return Transformer.from_crs(
        CRS.from_epsg(4326), load_crs_from_prj(prj_path), always_xy=True
    )


def point_to_segment_dist2(
    px: float, py: float, x1: float, y1: float, x2: float, y2: float
) -> float:
    dx, dy = x2 - x1, y2 - y1
    if dx == 0 and dy == 0:
        return (px - x1) ** 2 + (py - y1) ** 2
    t = ((px - x1) * dx + (py - y1) * dy) / (dx * dx + dy * dy)
    t = max(0, min(1, t))
    proj_x = x1 + t * dx
    proj_y = y1 + t * dy
    return (px - proj_x) ** 2 + (py - proj_y) ** 2


def point_to_polyline_dist2(px: float, py: float, points: List[Tuple[float, float]]) -> float:
    if not points:
        return float("inf")
    if len(points) == 1:
        return (px - points[0][0]) ** 2 + (py - points[0][1]) ** 2
    best = float("inf")
    for (x1, y1), (x2, y2) in zip(points, points[1:]):
        best = min(best, point_to_segment_dist2(px, py, x1, y1, x2, y2))
    return best


def wire_spec_at_point(
    lat: float,
    lon: float,
    prj_path: Path,
    shp_path: Path,
    transformer: Optional[Transformer] = None,
) -> Dict[str, Any]:
    """Return wire attributes for the line segment nearest to the given WGS84 point."""
    return wire_spec_between_points(lat, lon, lat, lon, prj_path, shp_path, transformer)


def wire_spec_between_points(
    lat1: float,
    lon1: float,
    lat2: float,
    lon2: float,
    prj_path: Path,
    shp_path: Path,
    transformer: Optional[Transformer] = None,
) -> Dict[str, Any]:
    """Return wire attributes for the line segment best matching the two WGS84 points."""
    if transformer is None:
        transformer = transformer_wgs84_to_layer(prj_path)
    x1, y1 = transformer.transform(lon1, lat1)
    x2, y2 = transformer.transform(lon2, lat2)

    reader = shapefile.Reader(str(shp_path))
    fields = [f[0] for f in reader.fields[1:]]
    idx = {name: fields.index(name) if name in fields else None for name in ("d_masterma", "d_neutralm", "d_orientat", "d_runtype")}

    best_i = None
    best_score = float("inf")
    best_d1 = best_d2 = None

    for i, (shape_rec, rec) in enumerate(zip(reader.shapes(), reader.records())):
        pts = shape_rec.points
        if not pts:
            continue
        d1_2 = point_to_polyline_dist2(x1, y1, pts)
        d2_2 = point_to_polyline_dist2(x2, y2, pts)
        score = max(d1_2, d2_2)
        if score < best_score:
            best_score = score
            best_i = i
            best_d1 = math.sqrt(d1_2)
            best_d2 = math.sqrt(d2_2)

    out = {
        "d_masterma": None,
        "d_neutralm": None,
        "d_orientat": None,
        "d_runtype": None,
        "dist1_ft": best_d1,
        "dist2_ft": best_d2,
        "line_index": best_i,
    }
    if best_i is not None:
        rec = reader.record(best_i)
        for k in out:
            if k in ("dist1_ft", "dist2_ft", "line_index"):
                continue
            if idx.get(k) is not None:
                out[k] = rec[idx[k]] or None
    return out


def wire_spec_between_points_oppd(
    lat1: float,
    lon1: float,
    lat2: float,
    lon2: float,
    base_path: Path,
    transformer: Optional[Transformer] = None,
) -> Dict[str, str]:
    """
    OPPD: Primary/Neutral from ElectricLine, Secondary from S_ElectricLine.
    Returns {Primary, Neutral, Secondary} (shapefile values).
    """
    prj = base_path / "ElectricLine selection.prj"
    el_shp = base_path / "ElectricLine selection.shp"
    s_shp = base_path / "S_ElectricLine selection.shp"
    out = {"Primary": "", "Neutral": "", "Secondary": ""}
    if not prj.exists() or not el_shp.exists():
        return out
    el_spec = wire_spec_between_points(
        lat1, lon1, lat2, lon2, prj, el_shp, transformer
    )
    out["Primary"] = (el_spec.get("d_masterma") or "").strip()
    out["Neutral"] = (el_spec.get("d_neutralm") or "").strip()
    if s_shp.exists():
        sec_spec = wire_spec_between_points(
            lat1, lon1, lat2, lon2, prj, s_shp, transformer
        )
        out["Secondary"] = (sec_spec.get("d_masterma") or "").strip()
    return out


def build_wire_spec_comparison(
    excel_path: Path,
    pplx_file_paths: List[str],
    shape_base_path: Path,
    extract_scid_fn,
    max_connections: Optional[int] = None,
) -> List[Dict[str, str]]:
    """
    Build comparison: Pole-Pole, Wire_Type, PPLX, Shape.
    Only when OPPD config. Returns list of dicts for change_log Wire Specs sheet.
    """
    nodes = load_nodes(excel_path)
    conns = load_connections_with_attrs(excel_path)
    if not conns or not nodes:
        return []
    if max_connections is not None:
        conns = conns[: max_connections]

    prj = shape_base_path / "ElectricLine selection.prj"
    if not prj.exists():
        return []
    trans = transformer_wgs84_to_layer(prj)

    scid_to_pplx: Dict[str, str] = {}
    for fp in pplx_file_paths:
        name = Path(fp).name
        scid = extract_scid_fn(name)
        base_scid = (scid.split() or [""])[0]
        if base_scid and base_scid not in scid_to_pplx:
            scid_to_pplx[base_scid] = fp

    LEN_TOLERANCE = 0.15  # 15% tolerance for span length match

    rows = []
    for c in conns:
        n1 = nodes.get(c["node_id_1"])
        n2 = nodes.get(c["node_id_2"])
        if not n1 or not n2:
            continue
        lat1 = n1.get("latitude")
        lon1 = n1.get("longitude")
        lat2 = n2.get("latitude")
        lon2 = n2.get("longitude")
        if lat1 is None or lon1 is None or lat2 is None or lon2 is None:
            continue
        try:
            lat1, lon1 = float(lat1), float(lon1)
            lat2, lon2 = float(lat2), float(lon2)
        except (TypeError, ValueError):
            continue
        scid1 = str(n1.get("scid") or "").strip()
        scid2 = str(n2.get("scid") or "").strip()
        base1 = (scid1.split() or [""])[0]
        base2 = (scid2.split() or [""])[0]
        if not base1 and not base2:
            continue
        pole_pair = f"{scid1 or base1}-{scid2 or base2}"

        shape_vals = wire_spec_between_points_oppd(
            lat1, lon1, lat2, lon2, shape_base_path, trans
        )

        span_ft = c.get("span_distance")
        span_in = (span_ft * 12) if span_ft is not None else None

        pplx_vals = {"Primary": "", "Neutral": "", "Secondary": ""}
        pplx_path = scid_to_pplx.get(base1) or scid_to_pplx.get(base2)
        if pplx_path:
            try:
                from src.core.handler import PPLXHandler

                h = PPLXHandler(pplx_path)
                by_type = h.get_spans_by_type_and_length()
                if span_in is not None:
                    for stype, lengths in by_type.items():
                        if stype not in pplx_vals:
                            continue
                        best_len = None
                        for ln in lengths:
                            if abs(ln - span_in) / max(span_in, 1) <= LEN_TOLERANCE:
                                if best_len is None or abs(ln - span_in) < abs(best_len - span_in):
                                    best_len = ln
                        if best_len is not None:
                            pplx_vals[stype] = lengths.get(best_len, "")
                else:
                    for stype in ("Primary", "Neutral", "Secondary"):
                        if stype in by_type and by_type[stype]:
                            first_len = min(by_type[stype])
                            pplx_vals[stype] = by_type[stype].get(first_len, "")
            except Exception:
                pass

        for wtype in ("Primary", "Neutral", "Secondary"):
            if shape_vals.get(wtype) or pplx_vals.get(wtype):
                rows.append({
                    "Pole-Pole": pole_pair,
                    "Wire_Type": wtype,
                    "PPLX": pplx_vals.get(wtype, ""),
                    "Shape": shape_vals.get(wtype, ""),
                })

    return rows


# ---------- Excel loading ----------
def load_nodes(excel_path: Path) -> Dict[str, Dict[str, Any]]:
    """node_id -> {scid, latitude, longitude, node_type, ...}. Only rows with lat/lon."""
    if not openpyxl or not excel_path.exists():
        return {}
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "nodes" not in wb.sheetnames:
        wb.close()
        return {}
    sheet = wb["nodes"]
    headers = {}
    for c in range(1, sheet.max_column + 1):
        h = sheet.cell(1, c).value
        if h:
            headers[h] = c
    need = ["node_id", "latitude", "longitude"]
    if any(h not in headers for h in need):
        wb.close()
        return {}
    data = {}
    for row in range(2, sheet.max_row + 1):
        nid = sheet.cell(row, headers["node_id"]).value
        lat = sheet.cell(row, headers["latitude"]).value
        lon = sheet.cell(row, headers["longitude"]).value
        if nid is None or lat is None or lon is None:
            continue
        try:
            lat, lon = float(lat), float(lon)
        except (TypeError, ValueError):
            continue
        row_data = {}
        for col_name, col_num in headers.items():
            v = sheet.cell(row, col_num).value
            row_data[col_name] = v
        data[str(nid)] = row_data
    wb.close()
    return data


def load_connections(excel_path: Path) -> List[Tuple[str, str]]:
    """List of (node_id_1, node_id_2)."""
    rows = load_connections_with_attrs(excel_path)
    return [(r["node_id_1"], r["node_id_2"]) for r in rows]


def load_connections_with_attrs(excel_path: Path) -> List[Dict[str, Any]]:
    """List of {node_id_1, node_id_2, span_distance}. span_distance in feet or None."""
    if not openpyxl or not excel_path.exists():
        return []
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "connections" not in wb.sheetnames:
        wb.close()
        return []
    sheet = wb["connections"]
    headers = {}
    for c in range(1, sheet.max_column + 1):
        h = sheet.cell(1, c).value
        if h:
            headers[h] = c
    if "node_id_1" not in headers or "node_id_2" not in headers:
        wb.close()
        return []
    span_col = headers.get("span_distance")
    out = []
    for row in range(2, sheet.max_row + 1):
        n1 = sheet.cell(row, headers["node_id_1"]).value
        n2 = sheet.cell(row, headers["node_id_2"]).value
        if n1 is None or n2 is None:
            continue
        span_dist = None
        if span_col:
            v = sheet.cell(row, span_col).value
            if v is not None:
                try:
                    span_dist = float(v)
                except (TypeError, ValueError):
                    pass
        out.append({
            "node_id_1": str(n1),
            "node_id_2": str(n2),
            "span_distance": span_dist,
        })
    wb.close()
    return out


def main() -> None:
    base = Path(__file__).parent
    excel_path = base / "Test Files" / "OPPD" / "NEOM104 Nodes-Sections-Connections XLSX.xlsx"
    prj_path = base / "data" / "OPPD" / "shape" / "ElectricLine selection.prj"
    shp_path = base / "data" / "OPPD" / "shape" / "ElectricLine selection.shp"

    if not excel_path.exists():
        print(f"Excel not found: {excel_path}")
        return
    if not shp_path.exists() or not prj_path.exists():
        print(f"Shapefile/PRJ not found: {shp_path} / {prj_path}")
        return

    nodes = load_nodes(excel_path)
    connections = load_connections(excel_path)
    transformer = transformer_wgs84_to_layer(prj_path)

    print(f"Nodes with coords: {len(nodes)}, Connections: {len(connections)}")
    print("Pole Number (SCID) 1->2, 2->3, etc.: wire spec between those poles\n")

    results: List[Tuple[str, str, Optional[str], Optional[str], float, float, Dict[str, Any]]] = []
    skipped = 0
    for nid1, nid2 in connections:
        n1 = nodes.get(nid1)
        n2 = nodes.get(nid2)
        if not n1 or not n2:
            skipped += 1
            continue
        lat1 = n1.get("latitude")
        lon1 = n1.get("longitude")
        lat2 = n2.get("latitude")
        lon2 = n2.get("longitude")
        if lat1 is None or lon1 is None or lat2 is None or lon2 is None:
            skipped += 1
            continue
        try:
            lat1, lon1 = float(lat1), float(lon1)
            lat2, lon2 = float(lat2), float(lon2)
        except (TypeError, ValueError):
            skipped += 1
            continue
        scid1 = n1.get("scid")
        scid2 = n2.get("scid")
        if scid1 is not None:
            scid1 = str(scid1).strip()
        if scid2 is not None:
            scid2 = str(scid2).strip()

        spec = wire_spec_between_points(lat1, lon1, lat2, lon2, prj_path, shp_path, transformer)
        results.append((scid1 or nid1, scid2 or nid2, nid1, nid2, lat1, lon1, spec))

    # Sort by (scid1, scid2) when numeric so we get 1→2, 2→3 order for display
    def sort_key(r):
        s1, s2 = r[0], r[1]
        try:
            a, b = int(s1), int(s2)
            return (0, a, b)
        except (ValueError, TypeError):
            return (1, str(s1), str(s2))

    results.sort(key=sort_key)

    print(f"Skipped (no coords): {skipped}\n")
    print(f"{'Pole (SCID)':<12} {'->':^4} {'Pole (SCID)':<12} {'d_masterma (primary)':<24} {'d_neutralm':<14} {'d_orientat':<8} {'dist1_ft':<10} {'dist2_ft':<10}")
    print("-" * 100)
    for scid1, scid2, _n1, _n2, _la1, _lo1, spec in results:
        master = (spec.get("d_masterma") or "").strip() or "-"
        neutral = (spec.get("d_neutralm") or "").strip() or "-"
        orient = (spec.get("d_orientat") or "").strip() or "-"
        d1 = spec.get("dist1_ft")
        d2 = spec.get("dist2_ft")
        d1s = f"{d1:.1f}" if d1 is not None else "-"
        d2s = f"{d2:.1f}" if d2 is not None else "-"
        print(f"{str(scid1):<12} {'->':^4} {str(scid2):<12} {master:<24} {neutral:<14} {orient:<8} {d1s:<10} {d2s:<10}")

    print("\nDone.")


if __name__ == "__main__":
    main()
