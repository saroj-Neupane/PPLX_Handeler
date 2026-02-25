"""Excel data loading for PPLX processing."""

import os
from typing import Callable, Dict, Optional

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def load_excel_data(
    excel_path: str,
    log_callback: Optional[Callable[[str], None]] = None,
) -> Dict[str, Dict]:
    """
    Load and filter Excel data based on node_type=pole and pole_status!=underground.
    Returns dict mapping SCID -> row data.
    """
    if log_callback:
        log_callback(f"Looking for Excel file at: {excel_path}")
        log_callback(f"Current working directory: {os.getcwd()}")

    if not excel_path or not os.path.exists(excel_path):
        if log_callback:
            log_callback(f"Excel Error: File not found or path empty: {excel_path}")
        return {}

    if not EXCEL_AVAILABLE:
        if log_callback:
            log_callback("Excel Support: openpyxl library not available.")
        return {}

    try:
        workbook = openpyxl.load_workbook(excel_path)
        if "nodes" not in workbook.sheetnames:
            if log_callback:
                log_callback("Excel Error: No 'nodes' sheet found")
            return {}

        sheet = workbook["nodes"]
        data = {}
        headers = {}
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header:
                headers[header.lower()] = col

        required_cols = ["scid", "node_type", "pole_status"]
        optional_cols = ["mr_note", "pole_tag_company", "pole_tag_tagtext"]
        missing_cols = [c for c in required_cols if c not in headers]
        if missing_cols:
            if log_callback:
                log_callback(f"Excel Error: Missing required columns: {missing_cols}")
            return {}

        if log_callback:
            missing_optional = [c for c in optional_cols if c not in headers]
            if missing_optional:
                log_callback(f"Excel Warning: Optional columns not found: {missing_optional}")

        valid_count = 0
        total_count = 0
        skipped_no_scid = 0
        skipped_filters = 0

        for row in range(2, sheet.max_row + 1):
            total_count += 1
            scid = sheet.cell(row=row, column=headers["scid"]).value
            node_type = sheet.cell(row=row, column=headers["node_type"]).value
            pole_status = sheet.cell(row=row, column=headers["pole_status"]).value

            if not scid:
                skipped_no_scid += 1
                continue

            if (
                node_type
                and str(node_type).lower() == "pole"
                and (pole_status is None or str(pole_status).lower() != "underground")
            ):
                row_data = {}
                for col_name, col_num in headers.items():
                    cell_value = sheet.cell(row=row, column=col_num).value
                    row_data[col_name] = (
                        str(cell_value) if cell_value is not None else ""
                    )
                data[str(scid)] = row_data
                valid_count += 1
            else:
                skipped_filters += 1

        workbook.close()
        if log_callback:
            log_callback(
                f"Excel loaded: {valid_count} valid pole entries "
                f"(total: {total_count}, skipped no SCID: {skipped_no_scid}, "
                f"skipped filters: {skipped_filters})"
            )
        return data

    except Exception as e:
        if log_callback:
            log_callback(f"Excel Error: {str(e)}")
        return {}
