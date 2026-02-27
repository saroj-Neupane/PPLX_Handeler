"""Change log Excel export, shared by GUI and headless processing."""

import re
from typing import Dict, List, Optional

from src.core.utils import leading_int

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

_AUX_HEADERS = [
    "File Name", "MR Note",
    "Aux Data 1", "Aux Data 2", "Aux Data 3", "Aux Data 4", "Aux Data 5",
]
_WIRE_HEADERS = ["Pole", "To Pole", "Wire_Type", "PPLX", "Shape", "QC"]
_SPANS_HEADERS = ["Pole", "To Pole", "Span Type", "Katapult", "PPLX", "QC"]


def _col_width(ws, col_idx: int, max_width: int = 60) -> float:
    """Estimate column width from the longest cell value in the column."""
    best = 0
    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value is None:
                continue
            # For multiline content take the longest line
            line_max = max(len(ln) for ln in str(cell.value).splitlines()) if str(cell.value) else 0
            best = max(best, line_max)
    return min(best + 4, max_width)


def _apply_table(ws, name: str) -> None:
    """Wrap the worksheet data in an Excel table with header filters."""
    max_col_letter = get_column_letter(ws.max_column)
    ref = f"A1:{max_col_letter}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _autofit(ws) -> None:
    """Set each column width based on its longest cell content."""
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _col_width(ws, col_idx)


def _normalize_ws(s: str) -> str:
    """Collapse whitespace for wire spec comparison."""
    return " ".join(s.split())


def write_change_log(
    path: str,
    csv_data: List[Dict],
    wire_spec_data: Optional[List[Dict]] = None,
    wire_spec_mapping: Optional[Dict[str, str]] = None,
    spans_data: Optional[List[Dict]] = None,
) -> bool:
    """
    Write change log xlsx with formatted Aux Data, Wire Specs, and Spans sheets.
    Returns True on success, False if openpyxl unavailable or no data.
    """
    if not OPENPYXL_AVAILABLE or not (csv_data or wire_spec_data or spans_data):
        return False

    wb = Workbook()
    wb.remove(wb.active)

    # --- Aux Data sheet ---
    ws_aux = wb.create_sheet("Aux Data", 0)
    for col, h in enumerate(_AUX_HEADERS, 1):
        ws_aux.cell(row=1, column=col, value=h)
    sorted_aux = sorted(csv_data or [], key=lambda r: leading_int(r.get("File Name", "")))
    for row_idx, row in enumerate(sorted_aux, start=2):
        for col, h in enumerate(_AUX_HEADERS, 1):
            ws_aux.cell(row=row_idx, column=col, value=row.get(h, ""))
    if ws_aux.max_row > 1:
        _apply_table(ws_aux, "AuxData")
    _autofit(ws_aux)

    # --- Wire Specs sheet ---
    ws_wire = wb.create_sheet("Wire Specs", 1)
    for col, h in enumerate(_WIRE_HEADERS, 1):
        ws_wire.cell(row=1, column=col, value=h)

    # Build normalized lookup from mapping (collapse whitespace, case-insensitive)
    norm_map: Dict[str, str] = {}
    if wire_spec_mapping:
        for k, v in wire_spec_mapping.items():
            key = _normalize_ws(k).lower()
            norm_map[key] = _normalize_ws(v).lower()

    wire_rows = []
    for row in (wire_spec_data or []):
        pole_pole = row.get("Pole-Pole", "")
        parts = pole_pole.split("-", 1)
        pplx = row.get("PPLX", "")
        shape = row.get("Shape", "")
        qc = ""
        if pplx and shape and norm_map:
            expected = norm_map.get(_normalize_ws(pplx).lower())
            if expected is not None:
                qc = "PASS" if expected == _normalize_ws(shape).lower() else "FAIL"
        wire_rows.append({
            "Pole": parts[0].strip() if parts else "",
            "To Pole": parts[1].strip() if len(parts) > 1 else "",
            "Wire_Type": row.get("Wire_Type", ""),
            "PPLX": pplx,
            "Shape": shape,
            "QC": qc,
        })
    wire_rows.sort(key=lambda r: (leading_int(r["Pole"]), leading_int(r["To Pole"])))

    qc_col_idx = _WIRE_HEADERS.index("QC") + 1
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    length_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    
    for row_idx, row in enumerate(wire_rows, start=2):
        for col, h in enumerate(_WIRE_HEADERS, 1):
            cell = ws_wire.cell(row=row_idx, column=col, value=row.get(h, ""))
            # Color-code QC column: PASS = light green, FAIL = light red
            if col == qc_col_idx:
                qc_value = str(row.get("QC", "")).strip().upper()
                if qc_value == "PASS":
                    cell.fill = pass_fill
                elif qc_value == "FAIL":
                    cell.fill = fail_fill
    if ws_wire.max_row > 1:
        _apply_table(ws_wire, "WireSpecs")
    _autofit(ws_wire)

    # --- Spans sheet (span count QC: Katapult vs PPLX per type) ---
    ws_spans = wb.create_sheet("Spans", 2)
    for col, h in enumerate(_SPANS_HEADERS, 1):
        ws_spans.cell(row=1, column=col, value=h)
    sorted_spans = sorted(spans_data or [], key=lambda r: (leading_int(r.get("Pole", "")), leading_int(r.get("To Pole", "")), r.get("Span Type", "")))
    qc_col_spans = _SPANS_HEADERS.index("QC") + 1
    for row_idx, row in enumerate(sorted_spans, start=2):
        for col, h in enumerate(_SPANS_HEADERS, 1):
            val = row.get(h, "")
            if h in ("Katapult", "PPLX") and isinstance(val, (int, float)):
                val = int(val)
            cell = ws_spans.cell(row=row_idx, column=col, value=val)
            if col == qc_col_spans:
                qc_val = str(val).strip().upper()
                if qc_val == "PASS":
                    cell.fill = pass_fill
                elif qc_val == "FAIL":
                    cell.fill = fail_fill
                elif qc_val == "LENGTH":
                    cell.fill = length_fill
    if ws_spans.max_row > 1:
        _apply_table(ws_spans, "SpansData")
    _autofit(ws_spans)

    wb.save(path)
    return True
