"""Processing frame for batch PPLX file processing."""

import os
import subprocess
import sys
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

from src.core.handler import PPLXHandler
from src.core.logic import (
    analyze_mr_note_for_aux_data,
    extract_scid_from_filename,
    clean_scid_keywords,
)
from src.core.utils import safe_filename_part, parse_keywords
from src.gui.constants import POLE_TAG_BLANK, THEME
from src.gui.frames.aux_data import _get_auto_fill_enabled

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Wire spec from shapefile (optional, OPPD config)
def _wire_spec_resolve_paths():
    base = Path(__file__).resolve().parent.parent.parent
    prj = base / "data" / "OPPD" / "shape" / "ElectricLine selection.prj"
    shp = base / "data" / "OPPD" / "shape" / "ElectricLine selection.shp"
    return prj, shp


def _wire_spec_base_path():
    """Base path for OPPD shapefiles (ElectricLine, S_ElectricLine)."""
    base = Path(__file__).resolve().parent.parent.parent
    return base / "data" / "OPPD" / "shape"


def _load_wire_spec_module():
    """Import wire spec helpers; ensure project root is on path so script is findable."""
    proj_root = Path(__file__).resolve().parent.parent.parent
    if str(proj_root) not in sys.path:
        sys.path.insert(0, str(proj_root))
    try:
        from wire_spec_from_excel import (
            wire_spec_at_point,
            transformer_wgs84_to_layer,
        )
        return True, wire_spec_at_point, transformer_wgs84_to_layer
    except ImportError:
        return False, None, None


WIRE_SPEC_AVAILABLE = False
wire_spec_at_point = None
transformer_wgs84_to_layer = None
build_wire_spec_comparison = None
_load = _load_wire_spec_module()
if _load[0]:
    WIRE_SPEC_AVAILABLE = True
    wire_spec_at_point = _load[1]
    transformer_wgs84_to_layer = _load[2]
    try:
        from wire_spec_from_excel import build_wire_spec_comparison
    except ImportError:
        build_wire_spec_comparison = None


def _set_aux_data_with_log(handler, aux_num: int, value: str, logs: list, prefix: str = "Set") -> bool:
    """Set aux data and append log. Returns success."""
    success = handler.set_aux_data(aux_num, value)
    action = "ERROR: Failed to set" if not success else prefix
    logs.append(f"  {action} Aux Data {aux_num}: {value}")
    return success


class ProcessingFrame(ttk.Frame):
    """Frame for processing controls and output."""

    def __init__(
        self,
        parent,
        config_manager,
        existing_frame,
        proposed_frame,
        aux_frame,
    ):
        super().__init__(parent)
        self.config_manager = config_manager
        self.existing_frame = existing_frame
        self.proposed_frame = proposed_frame
        self.aux_frame = aux_frame
        self.is_processing = False
        self.active_output_root = ""
        self.setup_ui()

    def setup_ui(self):
        ttk.Label(self, text="Process", font=("Segoe UI", 11, "bold")).pack(pady=(0, 6))
        ttk.Label(
            self,
            text="Output: Downloads/Processed PPLX (EXISTING + PROPOSED)",
            foreground=THEME["text_muted"],
            font=("Segoe UI", 9),
        ).pack(pady=(0, 6))
        self.process_button = ttk.Button(
            self, text="Process Files", command=self.start_processing
        )
        self.process_button.pack(pady=(0, 8))
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            self, variable=self.progress_var, maximum=100
        )
        self.progress_bar.pack(fill="x", pady=(0, 10))
        output_frame = ttk.LabelFrame(self, text="Processing Log", padding=5)
        output_frame.pack(fill="both", expand=True)
        self.output_text = scrolledtext.ScrolledText(
            output_frame,
            height=10,
            wrap=tk.WORD,
            bg=THEME["bg_card"],
            fg=THEME["text"],
            insertbackground=THEME["purple"],
            font=("Consolas", 10),
            relief="flat",
        )
        self.output_text.pack(fill="both", expand=True)

    def log_message(self, message: str):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.output_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.output_text.see(tk.END)
        self.output_text.update()

    def start_processing(self):
        if self.is_processing:
            return
        existing_files = self.existing_frame.get_files()
        proposed_files = self.proposed_frame.get_files()
        if not existing_files and not proposed_files:
            self.log_message("ERROR: Please select at least one folder containing PPLX files")
            return

        excel_path = self.aux_frame.get_excel_path()
        if not excel_path or not os.path.exists(excel_path):
            self.log_message("ERROR: Please select a valid Excel file before processing.")
            messagebox.showerror(
                "Excel Required",
                "Please select a valid Excel file in the Aux Data panel before processing.",
            )
            return

        aux_values = self.aux_frame.get_aux_values()
        if not aux_values:
            self.log_message(
                "WARNING: No Aux Data values specified. Files will be copied with auto-filled data only."
            )

        self.is_processing = True
        self.process_button.config(state="disabled", text="Processing...")
        self.output_text.delete(1.0, tk.END)

        category_data = [
            {
                "name": "EXISTING",
                "files": existing_files,
                "source_folder": self.existing_frame.get_source_path(),
            },
            {
                "name": "PROPOSED",
                "files": proposed_files,
                "source_folder": self.proposed_frame.get_source_path(),
            },
        ]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        downloads_dir = Path.home() / "Downloads"
        processed_base = downloads_dir / "Processed PPLX"
        processed_base.mkdir(parents=True, exist_ok=True)
        prefix = "O-Calcs"
        if excel_path:
            base_name = os.path.basename(excel_path)
            prefix_candidate = base_name.split(" ")[0].strip()
            if not prefix_candidate:
                prefix_candidate = Path(base_name).stem
            if prefix_candidate:
                prefix = f"{prefix_candidate}_O-Calcs"

        output_root = processed_base / f"{prefix}_{timestamp}"
        self.active_output_root = str(output_root)

        thread = threading.Thread(
            target=self.process_files,
            args=(category_data, aux_values, self.active_output_root),
        )
        thread.daemon = True
        thread.start()

    def process_files(
        self,
        category_data: List[Dict],
        aux_values: Dict[int, str],
        output_root: str,
    ):
        try:
            os.makedirs(output_root, exist_ok=True)
            excel_data = self.aux_frame.load_excel_data(log_callback=self.log_message)
            valid_scids = set(excel_data.keys()) if excel_data else set()

            total_files = sum(len(c["files"]) for c in category_data)
            if total_files == 0:
                self.log_message("No PPLX files found to process.")
                return

            self.log_message(f"Starting processing of {total_files} files across categories")
            self.log_message(f"Output root directory: {output_root}")
            if excel_data:
                self.log_message(f"{len(valid_scids)} non-underground poles found in Excel data")
            else:
                self.log_message("No Excel data loaded - processing all files")
            if aux_values:
                self.log_message(f"User Aux Data values (shared): {aux_values}")

            power_label = self.aux_frame.get_selected_config_power_label()
            if power_label != "POWER":
                self.log_message(f"Aux Data 4: replacing POWER with '{power_label}'")

            processed_count = 0
            summary: Dict[str, Dict] = {}

            for category in category_data:
                name = category["name"]
                files = category["files"]
                source_folder = category.get("source_folder") or ""
                output_dir = os.path.join(output_root, name)
                os.makedirs(output_dir, exist_ok=True)
                summary[name] = {
                    "successful": 0,
                    "failed": 0,
                    "skipped": 0,
                    "csv_path": "",
                    "output_dir": output_dir,
                }

                if not files:
                    self.log_message(f"\nCategory: {name} (no files selected, directory created)")
                    continue

                condition_value = name
                category_aux_values = {k: v for k, v in aux_values.items() if k != 3}
                csv_data = []
                wire_spec_data = []

                self.log_message(f"\nCategory: {name}")
                if source_folder:
                    self.log_message(f" Source: {source_folder}")
                self.log_message(f" Processing {len(files)} file{'s' if len(files) != 1 else ''}")

                ignore_keywords = getattr(
                    self.aux_frame, "ignore_scid_keywords", ""
                )
                auto_fill_aux1_enabled = _get_auto_fill_enabled(
                    self.aux_frame, 0, False
                )
                auto_fill_aux2_enabled = _get_auto_fill_enabled(
                    self.aux_frame, 1, False
                )

                def _get_str_var(name):
                    try:
                        var = getattr(self.aux_frame, name, None)
                        return (var.get() if var else "") or ""
                    except Exception:
                        return ""

                keyword_payload = {
                    "comm_keywords": parse_keywords(_get_str_var("comm_owners_var")),
                    "power_keywords": parse_keywords(_get_str_var("power_owners_var")),
                    "pco_keywords": parse_keywords(_get_str_var("pco_keywords_var")),
                    "aux5_keywords": parse_keywords(_get_str_var("aux5_keywords_var")),
                }

                def process_single_file(task):
                    index, file_path = task
                    logs = []
                    csv_row = None
                    status = "success"
                    filename = os.path.basename(file_path)

                    try:
                        scid = extract_scid_from_filename(filename)
                        pole_number = extract_scid_from_filename(filename)
                        clean_pole_number = clean_scid_keywords(
                            pole_number, ignore_keywords
                        )

                        if excel_data and scid not in valid_scids:
                            logs.append(
                                f"Skipping {filename}: SCID '{scid}' not found in Excel data"
                            )
                            return {
                                "index": index,
                                "status": "skipped",
                                "logs": logs,
                                "csv_row": None,
                            }

                        logs.append(
                            f"Processing: {filename} (SCID: {scid}, Pole Number: {pole_number} -> {clean_pole_number})"
                        )

                        handler = PPLXHandler(file_path)

                        if category_aux_values:
                            for aux_num, value in category_aux_values.items():
                                _set_aux_data_with_log(
                                    handler, aux_num, value, logs, "Set"
                                )

                        _set_aux_data_with_log(
                            handler, 3, condition_value, logs, "Auto-set"
                        )

                        pole_tag = POLE_TAG_BLANK
                        mr_note = ""

                        if excel_data and scid in excel_data:
                            row_data = excel_data[scid]

                            if auto_fill_aux1_enabled:
                                pole_owner = row_data.get("pole_tag_company", "")
                                if pole_owner:
                                    _set_aux_data_with_log(
                                        handler, 1, pole_owner, logs, "Auto-filled"
                                    )

                            if auto_fill_aux2_enabled:
                                excel_pole_tag = row_data.get(
                                    "pole_tag_tagtext", ""
                                ).strip()
                                pole_tag = (
                                    excel_pole_tag
                                    if excel_pole_tag
                                    else POLE_TAG_BLANK
                                )
                                _set_aux_data_with_log(
                                    handler, 2, pole_tag, logs, "Auto-filled"
                                )
                            else:
                                pole_tag = category_aux_values.get(
                                    2, POLE_TAG_BLANK
                                )
                                _set_aux_data_with_log(
                                    handler, 2, pole_tag, logs, "Set (manual)"
                                )

                            mr_note = row_data.get("mr_note", "")
                            aux_data_4, aux_data_5 = analyze_mr_note_for_aux_data(
                                mr_note,
                                comm_keywords=keyword_payload["comm_keywords"],
                                power_keywords=keyword_payload["power_keywords"],
                                pco_keywords=keyword_payload["pco_keywords"],
                                aux5_keywords=keyword_payload["aux5_keywords"],
                            )
                            if power_label != "POWER":
                                aux_data_4 = aux_data_4.replace(
                                    "POWER", power_label
                                )

                            _set_aux_data_with_log(
                                handler, 4, aux_data_4, logs, "Auto-filled"
                            )
                            if mr_note:
                                logs.append(
                                    f"    Based on mr_note: {mr_note[:50]}{'...' if len(mr_note) > 50 else ''}"
                                )
                            _set_aux_data_with_log(
                                handler, 5, aux_data_5, logs, "Auto-filled"
                            )
                        else:
                            pole_tag = (
                                category_aux_values.get(2, POLE_TAG_BLANK)
                                if not auto_fill_aux2_enabled
                                else pole_tag
                            )
                            _set_aux_data_with_log(
                                handler, 2, pole_tag, logs, "Set (fallback)"
                            )

                        if excel_data and scid in excel_data:
                            if auto_fill_aux2_enabled:
                                pole_tag = excel_data[scid].get(
                                    "pole_tag_tagtext", pole_tag
                                )
                            mr_note = excel_data[scid].get("mr_note", mr_note)

                        final_aux_data = handler.get_aux_data()
                        aux_data_4 = final_aux_data.get("Aux Data 4", "")
                        if aux_data_4 == "PCO":
                            clean_pole_number = f"{clean_pole_number} PCO"
                            logs.append(
                                f"  Aux Data 4 is 'PCO', appending to pole number: {clean_pole_number}"
                            )

                        clean_pole_number_safe = safe_filename_part(
                            clean_pole_number,
                            ". " if aux_data_4 == "PCO" else "",
                        )
                        clean_pole_tag = safe_filename_part(pole_tag, " ")
                        clean_condition = safe_filename_part(condition_value)

                        new_filename = f"{clean_pole_number_safe}_{clean_pole_tag}_{clean_condition}.pplx"
                        output_file = os.path.join(output_dir, new_filename)

                        handler.set_pole_attribute(
                            "Pole Number", clean_pole_number
                        )
                        logs.append(f"  Set Pole Number: {clean_pole_number}")

                        description_override = os.path.splitext(new_filename)[0]
                        handler.set_pole_attribute(
                            "DescriptionOverride", description_override
                        )
                        logs.append(
                            f"  Set DescriptionOverride: {description_override}"
                        )

                        handler.save_file(output_file)
                        logs.append(f"  Saved: {os.path.basename(output_file)}")

                        csv_row = {
                            "File Name": filename,
                            "MR Note": mr_note,
                            "Aux Data 1": final_aux_data.get(
                                "Aux Data 1", "Unset"
                            ),
                            "Aux Data 2": final_aux_data.get(
                                "Aux Data 2", "Unset"
                            ),
                            "Aux Data 3": final_aux_data.get(
                                "Aux Data 3", "Unset"
                            ),
                            "Aux Data 4": final_aux_data.get(
                                "Aux Data 4", "Unset"
                            ),
                            "Aux Data 5": final_aux_data.get(
                                "Aux Data 5", "Unset"
                            ),
                        }

                    except Exception as e:
                        logs.append(f"  Error processing {filename}: {str(e)}")
                        status = "failed"

                    return {
                        "index": index,
                        "status": status,
                        "logs": logs,
                        "csv_row": csv_row,
                    }

                max_workers = min(8, max(1, (os.cpu_count() or 1)))
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    for result in executor.map(
                        process_single_file, enumerate(files)
                    ):
                        for entry in result["logs"]:
                            self.log_message(entry)
                        status = result["status"]
                        if status == "success":
                            summary[name]["successful"] += 1
                            if result["csv_row"]:
                                csv_data.append(result["csv_row"])
                        elif status == "failed":
                            summary[name]["failed"] += 1
                        elif status == "skipped":
                            summary[name]["skipped"] += 1
                        processed_count += 1
                        self.progress_var.set(
                            (processed_count / total_files) * 100
                        )

                # OPPD config: build wire spec comparison (Pole-Pole, Wire_Type, PPLX, Shape)
                if (
                    build_wire_spec_comparison
                    and self.config_manager.config_name == "OPPD"
                    and files
                ):
                    excel_path = self.aux_frame.get_excel_path()
                    shape_base = _wire_spec_base_path()
                    if excel_path and shape_base.exists():
                        try:
                            wire_spec_data = build_wire_spec_comparison(
                                Path(excel_path),
                                files,
                                shape_base,
                                extract_scid_from_filename,
                            )
                            if wire_spec_data:
                                self.log_message(
                                    f"  Wire spec comparison: {len(wire_spec_data)} rows"
                                )
                        except Exception as e:
                            self.log_message(
                                f"  Wire spec comparison failed: {e}"
                            )

                if csv_data or wire_spec_data:
                    change_log_path = os.path.join(
                        output_dir, "change_log.xlsx"
                    )
                    try:
                        if OPENPYXL_AVAILABLE:
                            wb = Workbook()
                            wb.remove(wb.active)
                            # Sheet 1: Aux Data (as before)
                            ws_aux = wb.create_sheet("Aux Data", 0)
                            aux_headers = [
                                "File Name",
                                "MR Note",
                                "Aux Data 1",
                                "Aux Data 2",
                                "Aux Data 3",
                                "Aux Data 4",
                                "Aux Data 5",
                            ]
                            for col, h in enumerate(aux_headers, 1):
                                ws_aux.cell(row=1, column=col, value=h)
                            for row_idx, row_dict in enumerate(
                                csv_data, start=2
                            ):
                                for col, h in enumerate(aux_headers, 1):
                                    ws_aux.cell(
                                        row=row_idx,
                                        column=col,
                                        value=row_dict.get(h, ""),
                                    )
                            # Sheet 2: Wire Specs (Pole-Pole, Wire_Type, PPLX, Shape)
                            ws_wire = wb.create_sheet("Wire Specs", 1)
                            wire_headers = [
                                "Pole-Pole",
                                "Wire_Type",
                                "PPLX",
                                "Shape",
                            ]
                            for col, h in enumerate(wire_headers, 1):
                                ws_wire.cell(row=1, column=col, value=h)
                            for row_idx, row_dict in enumerate(
                                wire_spec_data, start=2
                            ):
                                for col, h in enumerate(wire_headers, 1):
                                    ws_wire.cell(
                                        row=row_idx,
                                        column=col,
                                        value=row_dict.get(h, ""),
                                    )
                            wb.save(change_log_path)
                            self.log_message(
                                f"{name} change log saved: {change_log_path}"
                            )
                            summary[name]["csv_path"] = change_log_path
                        else:
                            self.log_message(
                                "openpyxl not available; skipping change_log.xlsx"
                            )
                    except Exception as e:
                        self.log_message(
                            f"Error saving {name} change log: {str(e)}"
                        )

            self.progress_var.set(100)
            self.log_message("\nProcessing complete!")
            for name, stats in summary.items():
                self.log_message(
                    f"{name} - Successful: {stats['successful']}, "
                    f"Failed: {stats['failed']}, Skipped: {stats['skipped']}"
                )
                if stats.get("csv_path"):
                    self.log_message(f"{name} CSV: {stats['csv_path']}")
                self.log_message(f"{name} Output: {stats['output_dir']}")
            self.log_message(f"Output root directory: {output_root}")

            try:
                if os.name == "nt":
                    os.startfile(output_root)
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", output_root])
                else:
                    subprocess.Popen(["xdg-open", output_root])
            except Exception as open_err:
                self.log_message(
                    f"Info: Unable to open output folder: {open_err}"
                )

        except Exception as e:
            self.log_message(f"Critical error: {str(e)}")
            messagebox.showerror("Error", f"Processing failed: {str(e)}")
        finally:
            self.is_processing = False
            self.process_button.config(state="normal", text="Process Files")
