"""
Wire spec between pole pairs from Excel (nodes + connections) and OPPD shapefiles.
ElectricLine: Primary (d_masterma), Neutral (d_neutralm).
S_ElectricLine: Secondary (d_masterma).
"""
import math
import os
import pickle
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import shapefile  # pyshp
from pyproj import CRS, Transformer
from src.core.utils import leading_int

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pyogrio
    PYOGRIO_AVAILABLE = True
except ImportError:
    PYOGRIO_AVAILABLE = False

try:
    from rtree import index as rtree_index
    RTREE_AVAILABLE = True
except ImportError:
    RTREE_AVAILABLE = False


# ---------- Bearing (direction) from coordinates ----------
def bearing_rad_from_lat_lon(
    lat1_deg: float, lon1_deg: float, lat2_deg: float, lon2_deg: float
) -> float:
    """Bearing from point 1 to point 2 in radians, in [0, 2*pi). Uses spherical formula."""
    lat1 = math.radians(lat1_deg)
    lon1 = math.radians(lon1_deg)
    lat2 = math.radians(lat2_deg)
    lon2 = math.radians(lon2_deg)
    dlon = lon2 - lon1
    x = math.sin(dlon) * math.cos(lat2)
    y = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(lat2) * math.cos(dlon)
    theta = math.atan2(x, y)
    if theta < 0:
        theta += 2.0 * math.pi
    return theta


def _angle_diff_rad(a: float, b: float) -> float:
    """Smallest difference between two angles in [0, 2*pi), in radians."""
    a = a % (2.0 * math.pi)
    b = b % (2.0 * math.pi)
    d = abs(a - b)
    if d > math.pi:
        d = 2.0 * math.pi - d
    return d


def _base_pole_id(scid: str) -> str:
    """Normalize SCID like '013.A' -> '013' for filename-based lookups."""
    s = (scid or "").strip()
    if not s:
        return ""
    token = (s.split() or [""])[0]
    return token.split(".", 1)[0]


# ---------- Coordinate transform & distance ----------
def load_crs_from_prj(prj_path: Path) -> CRS:
    return CRS.from_wkt(prj_path.read_text())


def transformer_wgs84_to_layer(prj_path: Path) -> Transformer:
    return Transformer.from_crs(
        CRS.from_epsg(4326), load_crs_from_prj(prj_path), always_xy=True
    )


def point_to_segment_dist2(
    px: float, py: float, x1: float, y1: float, x2: float, y2: float
) -> float:
    dx, dy = x2 - x1, y2 - y1
    if dx == 0 and dy == 0:
        return (px - x1) ** 2 + (py - y1) ** 2
    t = ((px - x1) * dx + (py - y1) * dy) / (dx * dx + dy * dy)
    t = max(0, min(1, t))
    proj_x = x1 + t * dx
    proj_y = y1 + t * dy
    return (px - proj_x) ** 2 + (py - proj_y) ** 2


def _np_point_to_segments_dist2(px: float, py: float, seg_starts: np.ndarray, seg_ends: np.ndarray) -> float:
    """Vectorized: min squared distance from point (px,py) to a batch of segments.
    seg_starts, seg_ends: (N, 2) numpy arrays."""
    if len(seg_starts) == 0:
        return float("inf")
    dx = seg_ends[:, 0] - seg_starts[:, 0]
    dy = seg_ends[:, 1] - seg_starts[:, 1]
    len2 = dx * dx + dy * dy
    # t parameter clamped to [0, 1]
    # Avoid division by zero for degenerate segments
    safe_len2 = np.where(len2 == 0, 1.0, len2)
    t = ((px - seg_starts[:, 0]) * dx + (py - seg_starts[:, 1]) * dy) / safe_len2
    t = np.clip(t, 0.0, 1.0)
    # For degenerate segments, t doesn't matter; project to start
    t = np.where(len2 == 0, 0.0, t)
    proj_x = seg_starts[:, 0] + t * dx
    proj_y = seg_starts[:, 1] + t * dy
    d2 = (px - proj_x) ** 2 + (py - proj_y) ** 2
    return float(np.min(d2))


# Node types that may be shown as REF when they have no SCID (case-insensitive)
_REF_ALLOWED_NODE_TYPES = frozenset({"pole", "reference", "drop pole"})


def _build_scid_to_pplx_map(pplx_file_paths: List[str], extract_scid_fn) -> Dict[str, str]:
    """Map base SCID -> first matching PPLX file path."""
    mapping: Dict[str, str] = {}
    for fp in pplx_file_paths:
        name = Path(fp).name
        scid = extract_scid_fn(name)
        base = (scid.split() or [""])[0]
        if base and base not in mapping:
            mapping[base] = fp
    return mapping


def _pole_order(scid: str) -> Tuple[int, str]:
    """Sort key for pole SCID: (0, int) if numeric else (1, str) so 002 < 003."""
    s = (scid or "").strip()
    try:
        return (0, int(s.split()[0] if s else 0))
    except (ValueError, TypeError):
        return (1, s)


def _get_node_type(node: Dict[str, Any]) -> str:
    """Return node_type from node dict (case-insensitive key)."""
    for k, v in (node or {}).items():
        if str(k).strip().lower() == "node_type":
            return str(v or "").strip().lower()
    return ""


def point_to_polyline_dist2(px: float, py: float, points: List[Tuple[float, float]]) -> float:
    if not points:
        return float("inf")
    if len(points) == 1:
        return (px - points[0][0]) ** 2 + (py - points[0][1]) ** 2
    best = float("inf")
    for (x1, y1), (x2, y2) in zip(points, points[1:]):
        best = min(best, point_to_segment_dist2(px, py, x1, y1, x2, y2))
    return best


# ---------- Cached shapefile layer ----------
_SEARCH_MARGIN = 500  # projected units (feet), expanded via fallback if needed
_SHPCACHE_DIR = ".cache"  # subdir next to .shp for persistent pickle cache

_ATTR_NAMES = ("d_masterma", "d_neutralm", "d_orientat", "d_runtype")


def _load_layer_data_pyogrio(shp_path: Path) -> Optional[Dict[str, Any]]:
    """Load shapefile with pyogrio (faster). Returns dict with points, records, bboxes, field_idx or None."""
    if not PYOGRIO_AVAILABLE:
        return None
    try:
        df = pyogrio.read_dataframe(str(shp_path))
    except Exception:
        return None
    if df is None or len(df) == 0 or "geometry" not in df.columns:
        return None
    points = []
    records = []
    raw_bboxes = []
    for idx, row in df.iterrows():
        geom = row["geometry"]
        if geom is None or getattr(geom, "is_empty", True):
            continue
        if hasattr(geom, "coords"):
            pts = list(geom.coords)
        elif hasattr(geom, "exterior"):
            pts = list(geom.exterior.coords)
        else:
            continue
        if not pts:
            continue
        points.append(pts)
        raw_bboxes.append(geom.bounds)
        rec = tuple(row.get(name) for name in _ATTR_NAMES)
        records.append(rec)
    if not points:
        return None
    field_idx = {name: i for i, name in enumerate(_ATTR_NAMES)}
    bboxes = np.array(raw_bboxes, dtype=np.float64)
    return {"points": points, "records": records, "bboxes": bboxes, "field_idx": field_idx}


class _ShapefileLayer:
    """Pre-loaded shapefile with numpy bbox and optional R-tree for fast spatial queries."""
    __slots__ = ("field_idx", "points", "records", "bboxes", "_rtree_idx", "_seg_starts", "_seg_ends")

    def __init__(self, shp_path: Path):
        self._rtree_idx = None
        self._seg_starts = []
        self._seg_ends = []
        data = _load_layer_data_pyogrio(Path(shp_path)) if PYOGRIO_AVAILABLE else None
        if data is not None:
            self.points = data["points"]
            self.records = data["records"]
            self.bboxes = data["bboxes"]
            self.field_idx = data["field_idx"]
            self._precompute_segments()
            self._build_rtree()
            return
        # Fallback: pyshp
        reader = shapefile.Reader(str(shp_path))
        fields = [f[0] for f in reader.fields[1:]]
        self.field_idx = {
            name: fields.index(name) if name in fields else None
            for name in _ATTR_NAMES
        }
        self.points = []
        self.records = []
        raw_bboxes = []
        for sr, rec in zip(reader.shapes(), reader.records()):
            if not sr.points:
                continue
            self.points.append(sr.points)
            self.records.append(rec)
            raw_bboxes.append(sr.bbox)  # [xmin, ymin, xmax, ymax]
        self.bboxes = np.array(raw_bboxes, dtype=np.float64) if raw_bboxes else np.empty((0, 4))
        self._precompute_segments()
        self._build_rtree()

    def _precompute_segments(self) -> None:
        """Precompute numpy segment arrays for each polyline for vectorized distance calc."""
        self._seg_starts = []
        self._seg_ends = []
        for pts in self.points:
            if len(pts) < 2:
                # Single-point: create a degenerate segment
                arr = np.array(pts, dtype=np.float64)
                self._seg_starts.append(arr)
                self._seg_ends.append(arr)
            else:
                arr = np.array(pts, dtype=np.float64)
                self._seg_starts.append(arr[:-1])
                self._seg_ends.append(arr[1:])

    def _build_rtree(self) -> None:
        """Build R-tree index from bboxes for O(log n) candidate lookup (when rtree is installed)."""
        if not RTREE_AVAILABLE or len(self.bboxes) == 0:
            return
        try:
            idx = rtree_index.Index(interleaved=True)
            for i in range(len(self.bboxes)):
                b = self.bboxes[i]
                idx.insert(i, (float(b[0]), float(b[1]), float(b[2]), float(b[3])))
            self._rtree_idx = idx
        except Exception:
            self._rtree_idx = None

    @classmethod
    def from_cache(cls, cache_path: Path) -> "_ShapefileLayer":
        """Load layer from a previously saved pickle cache."""
        with open(cache_path, "rb") as f:
            data = pickle.load(f)
        self = cls.__new__(cls)
        self.points = data["points"]
        self.records = data["records"]
        self.bboxes = data["bboxes"]
        self.field_idx = data["field_idx"]
        self._rtree_idx = None
        self._seg_starts = []
        self._seg_ends = []
        self._precompute_segments()
        self._build_rtree()
        return self

    def _to_cache_data(self) -> Dict[str, Any]:
        """Data dict suitable for pickle (for disk cache)."""
        return {
            "points": self.points,
            "records": self.records,
            "bboxes": self.bboxes,
            "field_idx": self.field_idx,
        }

    def query(self, x1, y1, x2, y2) -> Dict[str, Any]:
        """Find best matching line for the two projected points."""
        best_i, best_score, best_d1_2, best_d2_2 = None, float("inf"), 0.0, 0.0

        if len(self.bboxes) == 0:
            return self._build_result(None, 0.0, 0.0)

        lo_x = min(x1, x2) - _SEARCH_MARGIN
        lo_y = min(y1, y2) - _SEARCH_MARGIN
        hi_x = max(x1, x2) + _SEARCH_MARGIN
        hi_y = max(y1, y2) + _SEARCH_MARGIN

        # R-tree: O(log n) candidate lookup when available; else vectorized bbox scan
        if self._rtree_idx is not None:
            try:
                candidates = np.array(list(self._rtree_idx.intersection((lo_x, lo_y, hi_x, hi_y))), dtype=np.intp)
            except Exception:
                candidates = np.where(
                    (self.bboxes[:, 2] >= lo_x) & (self.bboxes[:, 0] <= hi_x) &
                    (self.bboxes[:, 3] >= lo_y) & (self.bboxes[:, 1] <= hi_y)
                )[0]
        else:
            mask = (
                (self.bboxes[:, 2] >= lo_x) & (self.bboxes[:, 0] <= hi_x) &
                (self.bboxes[:, 3] >= lo_y) & (self.bboxes[:, 1] <= hi_y)
            )
            candidates = np.where(mask)[0]

        if len(candidates) == 0:
            candidates = np.arange(len(self.points))

        # Vectorized distance computation using precomputed segment arrays
        for i in candidates:
            starts = self._seg_starts[i]
            ends = self._seg_ends[i]
            d1_2 = _np_point_to_segments_dist2(x1, y1, starts, ends)
            d2_2 = _np_point_to_segments_dist2(x2, y2, starts, ends)
            score = max(d1_2, d2_2)
            if score < best_score:
                best_score = score
                best_i = int(i)
                best_d1_2 = d1_2
                best_d2_2 = d2_2

        return self._build_result(best_i, best_d1_2, best_d2_2)

    def _build_result(self, best_i, best_d1_2, best_d2_2) -> Dict[str, Any]:
        out: Dict[str, Any] = {
            "d_masterma": None, "d_neutralm": None,
            "d_orientat": None, "d_runtype": None,
            "dist1_ft": None, "dist2_ft": None, "line_index": best_i,
        }
        if best_i is not None:
            out["dist1_ft"] = math.sqrt(best_d1_2)
            out["dist2_ft"] = math.sqrt(best_d2_2)
            rec = self.records[best_i]
            for k, idx in self.field_idx.items():
                if idx is not None:
                    out[k] = rec[idx] or None
        return out


# In-memory cache: path string -> _ShapefileLayer (per process)
_layer_cache: Dict[str, _ShapefileLayer] = {}


def _get_layer(shp_path: Path, log_callback=None) -> _ShapefileLayer:
    """Load shapefile layer, using in-memory cache and optional disk cache (.cache/*.pkl)."""
    key = str(shp_path)
    if key in _layer_cache:
        return _layer_cache[key]

    shp_path = Path(shp_path)
    cache_dir = shp_path.parent / _SHPCACHE_DIR
    cache_path = cache_dir / f"{shp_path.stem}.pkl"

    # Use disk cache if it exists and is newer than the .shp
    try:
        if cache_path.exists() and shp_path.exists():
            shp_mtime = os.path.getmtime(shp_path)
            cache_mtime = os.path.getmtime(cache_path)
            if cache_mtime >= shp_mtime:
                layer = _ShapefileLayer.from_cache(cache_path)
                _layer_cache[key] = layer
                if log_callback:
                    log_callback(f"  Wire spec: using disk cache for {shp_path.name}")
                return layer
    except (OSError, pickle.PickleError):
        pass

    # Load from .shp and optionally write disk cache
    if log_callback:
        log_callback(f"  Wire spec: loading {shp_path.name} from .shp (writing disk cache)")
    layer = _ShapefileLayer(shp_path)
    _layer_cache[key] = layer
    try:
        cache_dir.mkdir(parents=True, exist_ok=True)
        with open(cache_path, "wb") as f:
            pickle.dump(layer._to_cache_data(), f, protocol=pickle.HIGHEST_PROTOCOL)
    except OSError:
        pass  # ignore cache write errors (e.g. read-only dir)
    return layer


# ---------- Public query functions ----------
def wire_spec_at_point(
    lat: float,
    lon: float,
    prj_path: Path,
    shp_path: Path,
    transformer: Optional[Transformer] = None,
) -> Dict[str, Any]:
    """Return wire attributes for the line segment nearest to the given WGS84 point."""
    return wire_spec_between_points(lat, lon, lat, lon, prj_path, shp_path, transformer)


def wire_spec_between_points(
    lat1: float,
    lon1: float,
    lat2: float,
    lon2: float,
    prj_path: Path,
    shp_path: Path,
    transformer: Optional[Transformer] = None,
) -> Dict[str, Any]:
    """Return wire attributes for the line segment best matching the two WGS84 points."""
    if transformer is None:
        transformer = transformer_wgs84_to_layer(prj_path)
    x1, y1 = transformer.transform(lon1, lat1)
    x2, y2 = transformer.transform(lon2, lat2)

    layer = _get_layer(shp_path)
    return layer.query(x1, y1, x2, y2)


def wire_spec_between_points_oppd(
    lat1: float,
    lon1: float,
    lat2: float,
    lon2: float,
    base_path: Path,
    transformer: Optional[Transformer] = None,
) -> Dict[str, str]:
    """
    OPPD: Primary/Neutral from ElectricLine, Secondary from S_ElectricLine.
    Returns {Primary, Neutral, Secondary} (shapefile values).
    """
    prj = base_path / "ElectricLine selection.prj"
    el_shp = base_path / "ElectricLine selection.shp"
    s_shp = base_path / "S_ElectricLine selection.shp"
    out = {"Primary": "", "Neutral": "", "Secondary": ""}
    if not prj.exists() or not el_shp.exists():
        return out
    el_spec = wire_spec_between_points(
        lat1, lon1, lat2, lon2, prj, el_shp, transformer
    )
    out["Primary"] = (el_spec.get("d_masterma") or "").strip()
    out["Neutral"] = (el_spec.get("d_neutralm") or "").strip()
    if s_shp.exists():
        sec_spec = wire_spec_between_points(
            lat1, lon1, lat2, lon2, prj, s_shp, transformer
        )
        out["Secondary"] = (sec_spec.get("d_masterma") or "").strip()
    return out


def build_wire_spec_comparison(
    excel_path: Path,
    pplx_file_paths: List[str],
    shape_base_path: Path,
    extract_scid_fn,
    max_connections: Optional[int] = None,
    log_callback=None,
    pplx_cache: Optional[Dict[str, Any]] = None,
    nodes: Optional[Dict[str, Dict[str, Any]]] = None,
    conns: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, str]]:
    """
    Build comparison: Pole-Pole, Wire_Type, PPLX, Shape.
    Only when OPPD config. Returns list of dicts for change_log Wire Specs sheet.
    If pplx_cache is provided (path -> PPLXHandler), it is used and updated for missing paths.
    Pass nodes/conns to avoid redundant Excel loading (loaded once, shared with spans build).
    """
    _log = log_callback or (lambda msg: None)

    if nodes is None:
        nodes = load_nodes(excel_path)
    if conns is None:
        conns = load_connections_with_attrs(excel_path)
    if not conns or not nodes:
        return []
    if max_connections is not None:
        conns = conns[: max_connections]

    _log(f"  Wire spec: {len(conns)} connections, {len(nodes)} nodes")

    prj = shape_base_path / "ElectricLine selection.prj"
    if not prj.exists():
        _log(f"  Wire spec: PRJ file not found at {prj}")
        return []
    t0 = time.perf_counter()
    _log("  Wire spec: loading shapefiles...")
    trans = transformer_wgs84_to_layer(prj)
    el_shp = shape_base_path / "ElectricLine selection.shp"
    s_shp = shape_base_path / "S_ElectricLine selection.shp"
    el_layer = _get_layer(el_shp, log_callback=_log) if el_shp.exists() else None
    s_layer = _get_layer(s_shp, log_callback=_log) if s_shp.exists() else None
    t_load = time.perf_counter() - t0
    _log(f"  Wire spec: shapefiles loaded in {t_load:.1f}s" + (" (pyogrio)" if PYOGRIO_AVAILABLE else " (pyshp)"))
    if not el_layer:
        _log("  Wire spec: ElectricLine shapefile not found")
        return []

    t1 = time.perf_counter()
    projected: Dict[str, Tuple[float, float]] = {}
    for nid, nd in nodes.items():
        lat = nd.get("latitude")
        lon = nd.get("longitude")
        if lat is None or lon is None:
            continue
        try:
            x, y = trans.transform(float(lon), float(lat))
            projected[nid] = (x, y)
        except (TypeError, ValueError):
            continue
    _log(f"  Wire spec: {len(projected)} nodes projected")

    scid_to_pplx = _build_scid_to_pplx_map(pplx_file_paths, extract_scid_fn)

    LEN_TOLERANCE = 0.15  # 15% tolerance for span length match

    from src.core.handler import PPLXHandler
    if pplx_cache is None:
        pplx_cache = {}
    # One get_spans_by_type_and_length() per path (not per connection)
    wire_spec_by_path: Dict[str, Dict[str, Dict[float, str]]] = {}

    rows = []
    total = len(conns)
    t_loop = time.perf_counter()
    for ci, c in enumerate(conns):
        if ci > 0 and ci % 50 == 0:
            _log(f"  Wire spec: {ci}/{total} connections processed...")

        nid1, nid2 = c["node_id_1"], c["node_id_2"]
        xy1 = projected.get(nid1)
        xy2 = projected.get(nid2)
        if not xy1 or not xy2:
            continue

        n1 = nodes[nid1]
        n2 = nodes[nid2]
        scid1 = str(n1.get("scid") or "").strip()
        scid2 = str(n2.get("scid") or "").strip()
        base1 = (scid1.split() or [""])[0]
        base2 = (scid2.split() or [""])[0]
        if not base1 and not base2:
            continue
        # Only treat as REF when node without SCID is pole, reference, or drop pole
        if not base1 and _get_node_type(n1) not in _REF_ALLOWED_NODE_TYPES:
            continue
        if not base2 and _get_node_type(n2) not in _REF_ALLOWED_NODE_TYPES:
            continue
        # Use REF for node without SCID so change log shows e.g. Pole=001, To Pole=REF
        s1 = (scid1 or base1) or "REF"
        s2 = (scid2 or base2) or "REF"
        # Normalize to ascending order so change log shows 002-->003, 003-->004; REF last
        if _pole_order(s1) <= _pole_order(s2):
            pole_pair = f"{s1}-{s2}"
        else:
            pole_pair = f"{s2}-{s1}"

        # Query layers directly with pre-transformed coordinates
        el_result = el_layer.query(xy1[0], xy1[1], xy2[0], xy2[1])
        shape_vals = {
            "Primary": (el_result.get("d_masterma") or "").strip(),
            "Neutral": (el_result.get("d_neutralm") or "").strip(),
            "Secondary": "",
        }
        if s_layer:
            s_result = s_layer.query(xy1[0], xy1[1], xy2[0], xy2[1])
            shape_vals["Secondary"] = (s_result.get("d_masterma") or "").strip()

        span_ft = c.get("span_distance")
        span_in = (span_ft * 12) if span_ft is not None else None

        pplx_vals = {"Primary": "", "Neutral": "", "Secondary": ""}
        pplx_path = scid_to_pplx.get(base1) or scid_to_pplx.get(base2)
        if pplx_path:
            try:
                if pplx_path not in pplx_cache:
                    pplx_cache[pplx_path] = PPLXHandler(pplx_path)
                h = pplx_cache[pplx_path]
                if pplx_path not in wire_spec_by_path:
                    wire_spec_by_path[pplx_path] = h.get_spans_by_type_and_length()
                by_type = wire_spec_by_path[pplx_path]
                if span_in is not None:
                    for stype, lengths in by_type.items():
                        if stype not in pplx_vals:
                            continue
                        best_len = None
                        for ln in lengths:
                            if abs(ln - span_in) / max(span_in, 1) <= LEN_TOLERANCE:
                                if best_len is None or abs(ln - span_in) < abs(best_len - span_in):
                                    best_len = ln
                        if best_len is not None:
                            pplx_vals[stype] = lengths.get(best_len, "")
                else:
                    for stype in ("Primary", "Neutral", "Secondary"):
                        if stype in by_type and by_type[stype]:
                            first_len = min(by_type[stype])
                            pplx_vals[stype] = by_type[stype].get(first_len, "")
            except Exception:
                pass

        for wtype in ("Primary", "Neutral", "Secondary"):
            if shape_vals.get(wtype) or pplx_vals.get(wtype):
                rows.append({
                    "Pole-Pole": pole_pair,
                    "Wire_Type": wtype,
                    "PPLX": pplx_vals.get(wtype, ""),
                    "Shape": shape_vals.get(wtype, ""),
                })

    t_loop_elapsed = time.perf_counter() - t_loop
    _log(f"  Wire spec: {total} connections in {t_loop_elapsed:.1f}s (PPLX files loaded: {len(pplx_cache)})")
    return rows


# ---------- Excel loading ----------

def _extract_nodes_from_sheet(sheet, headers) -> Dict[str, Dict[str, Any]]:
    """Extract nodes data from an already-open worksheet."""
    need = ["node_id", "latitude", "longitude"]
    if any(h not in headers for h in need):
        return {}
    data = {}
    for row in sheet.iter_rows(min_row=2, values_only=False):
        cells = {h: row[c - 1].value for h, c in headers.items()}
        nid = cells.get("node_id")
        lat = cells.get("latitude")
        lon = cells.get("longitude")
        if nid is None or lat is None or lon is None:
            continue
        try:
            lat, lon = float(lat), float(lon)
        except (TypeError, ValueError):
            continue
        row_data = {h: cells.get(h) for h in headers}
        data[str(nid)] = row_data
    return data


def _extract_connections_from_sheet(sheet, headers) -> List[Dict[str, Any]]:
    """Extract connections data from an already-open worksheet."""
    if "node_id_1" not in headers or "node_id_2" not in headers:
        return []
    span_col = headers.get("span_distance")
    conn_id_col = headers.get("connection_id")
    out = []
    n1_idx = headers["node_id_1"] - 1
    n2_idx = headers["node_id_2"] - 1
    span_idx = (span_col - 1) if span_col else None
    conn_id_idx = (conn_id_col - 1) if conn_id_col else None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        n1 = row[n1_idx] if n1_idx < len(row) else None
        n2 = row[n2_idx] if n2_idx < len(row) else None
        if n1 is None or n2 is None:
            continue
        span_dist = None
        if span_idx is not None and span_idx < len(row):
            v = row[span_idx]
            if v is not None:
                try:
                    span_dist = float(v)
                except (TypeError, ValueError):
                    pass
        conn_id = None
        if conn_id_idx is not None and conn_id_idx < len(row):
            conn_id = row[conn_id_idx]
        out.append({
            "connection_id": str(conn_id) if conn_id is not None else None,
            "node_id_1": str(n1),
            "node_id_2": str(n2),
            "span_distance": span_dist,
        })
    return out


def _extract_sections_from_sheet(sheet, headers) -> List[Dict[str, Any]]:
    """
    Extract sections data from an already-open worksheet.
    Deduplicates rows with identical (connection_id, POAs) to avoid counting duplicates.
    """
    conn_id_col = headers.get("connection_id")
    if not conn_id_col:
        return []
    conn_id_idx = conn_id_col - 1
    poa_indices = []
    for i in range(1, 7):
        col = headers.get(f"POA_{i}")
        if col:
            poa_indices.append(col - 1)

    # Use a set to track unique (connection_id, POAs tuple) to deduplicate
    seen: set[tuple[str, tuple[str, ...]]] = set()
    out = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        conn_id = row[conn_id_idx] if conn_id_idx < len(row) else None
        if conn_id is None:
            continue
        poas = []
        for idx in poa_indices:
            if idx < len(row):
                v = row[idx]
                if v is not None and str(v).strip():
                    poas.append(str(v).strip())

        conn_id_str = str(conn_id)
        poas_tuple = tuple(poas)

        # Skip duplicate (connection_id, POAs) combinations
        key = (conn_id_str, poas_tuple)
        if key in seen:
            continue
        seen.add(key)

        out.append({"connection_id": conn_id_str, "POAs": poas})
    return out


def _get_sheet_headers(sheet) -> Dict[str, int]:
    """Extract column headers from first row. Returns {header_name: 1-based col index}."""
    headers = {}
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=False):
        for cell in row:
            if cell.value:
                headers[cell.value] = cell.column
        break
    return headers


def load_all_excel_data(excel_path: Path) -> Tuple[Dict[str, Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Load nodes, connections, and sections from a single workbook open.

    Returns (nodes, connections, sections). Opens the workbook once — much faster
    than calling load_nodes + load_connections_with_attrs + load_sections separately.
    """
    empty = ({}, [], [])
    if not openpyxl or not excel_path.exists():
        return empty
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        nodes = {}
        conns = []
        sections = []
        if "nodes" in wb.sheetnames:
            sheet = wb["nodes"]
            headers = _get_sheet_headers(sheet)
            nodes = _extract_nodes_from_sheet(sheet, headers)
        if "connections" in wb.sheetnames:
            sheet = wb["connections"]
            headers = _get_sheet_headers(sheet)
            conns = _extract_connections_from_sheet(sheet, headers)
        if "sections" in wb.sheetnames:
            sheet = wb["sections"]
            headers = _get_sheet_headers(sheet)
            sections = _extract_sections_from_sheet(sheet, headers)
        return nodes, conns, sections
    finally:
        wb.close()


def load_nodes(excel_path: Path) -> Dict[str, Dict[str, Any]]:
    """node_id -> {scid, latitude, longitude, node_type, ...}. Only rows with lat/lon."""
    nodes, _, _ = load_all_excel_data(excel_path)
    return nodes


def load_connections(excel_path: Path) -> List[Tuple[str, str]]:
    """List of (node_id_1, node_id_2)."""
    rows = load_connections_with_attrs(excel_path)
    return [(r["node_id_1"], r["node_id_2"]) for r in rows]


def load_connections_with_attrs(excel_path: Path) -> List[Dict[str, Any]]:
    """List of {connection_id, node_id_1, node_id_2, span_distance}."""
    _, conns, _ = load_all_excel_data(excel_path)
    return conns


def load_sections(excel_path: Path) -> List[Dict[str, Any]]:
    """List of {connection_id, POAs}. One row per connection."""
    _, _, sections = load_all_excel_data(excel_path)
    return sections


def load_midspan_heights_counts(
    excel_path: Path,
    log_callback=None,
    power_label: str = "OPPD",
) -> Dict[Tuple[str, str], int]:
    """
    Load Node and Midspan Heights workbook and derive approximate comm attachment counts
    per span pair: (base_scid_1, base_scid_2) -> count of non-power attachments.

    Assumptions (based on NEOM104 Node and Midspan Heights layout):
    - One sheet per SCID, named 'SCID XXX'.
    - Row 1: header with first cell 'SCID XXX', second+ cells unused.
    - Row 2: for each midspan column, text like 'midspan to 074_in_feet'.
    - Rows 3+: first column = company/owner, each midspan column cell = height string
      when that attachment participates in that span.
    - Non-power attachments are those whose company != power_label.
    """
    counts: Dict[Tuple[str, str], int] = {}
    if not openpyxl or not excel_path.exists():
        return counts
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        if log_callback:
            log_callback(f"  Midspan heights: failed to open {excel_path}: {e}")
        return counts

    for name in wb.sheetnames:
        name_str = str(name or "")
        if not name_str.startswith("SCID"):
            continue
        base_scid = name_str.replace("SCID", "").strip()
        if not base_scid:
            continue
        ws = wb[name]
        # Row 2 describes midspan columns: "midspan to 074_in_feet"
        row2_vals = [c.value for c in ws[2]]
        dest_cols: List[Tuple[int, str]] = []
        for col_idx, v in enumerate(row2_vals[1:], start=1):
            if not isinstance(v, str):
                continue
            m = re.search(r"midspan to\s+([0-9A-Za-z\.]+)", v)
            if not m:
                continue
            dest = m.group(1).strip()
            dest_base = (dest.split() or [""])[0]
            if not dest_base:
                continue
            dest_cols.append((col_idx, dest_base))
        if not dest_cols:
            continue

        # Rows 3+: each non-empty row with non-power company and non-empty height counts as 1 attachment
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not any(row):
                continue
            company = str(row[0] or "").strip()
            if not company:
                continue
            if power_label and company.upper() == power_label.upper():
                # Treat power owner as non-comm; skip
                continue
            for col_idx, dest_base in dest_cols:
                if col_idx >= len(row):
                    continue
                height_val = row[col_idx]
                if height_val is None or (isinstance(height_val, str) and not height_val.strip()):
                    continue
                key = tuple(sorted([base_scid, dest_base]))
                counts[key] = counts.get(key, 0) + 1

    wb.close()
    if log_callback and counts:
        log_callback(f"  Midspan heights: derived comm attachment counts for {len(counts)} span pairs")
    return counts


def build_spans_comparison_data(
    excel_path: Path,
    pplx_file_paths: List[str],
    extract_scid_fn,
    length_tolerance_in: float = 36.0,
    log_callback=None,
    span_type_mapping: Optional[Dict[str, Optional[str]]] = None,
    pplx_cache: Optional[Dict[str, Any]] = None,
    midspan_heights_path: Optional[Path] = None,
    nodes: Optional[Dict[str, Dict[str, Any]]] = None,
    conns: Optional[List[Dict[str, Any]]] = None,
    sections: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    """
    Compare span counts per type between Nodes-Sections-Connections (Katapult) and PPLX.
    Matching is by length: Katapult uses Connections.span_distance; PPLX uses SpanDistanceInInches.
    Returns list of dicts: Pole, To Pole, Span Type, Katapult, PPLX, QC.

    Args:
        length_tolerance_in: Absolute tolerance in inches for matching span lengths (default: 36\" = 3').
        span_type_mapping: Dict mapping Katapult POA values (lowercase) to canonical SpanType
                          (lowercase) or None to exclude. If None, loads from config.
        nodes: Pre-loaded nodes data (avoids redundant Excel load).
        conns: Pre-loaded connections data (avoids redundant Excel load).
        sections: Pre-loaded sections data (avoids redundant Excel load).
    """
    _log = log_callback or (lambda msg: None)

    # Load span_type_mapping from config if not provided
    if span_type_mapping is None:
        from src.config.manager import get_active_config_name
        from pathlib import Path as PathLib
        import json
        root = PathLib(__file__).resolve().parents[2]
        config_name = get_active_config_name()
        if not config_name.endswith(".json"):
            config_name = f"{config_name}.json"
        config_path = root / "config" / config_name
        span_type_mapping = {}
        try:
            if config_path.exists():
                with open(config_path, "r") as f:
                    config = json.load(f)
                span_type_mapping = config.get("span_type_mapping", {})
                span_type_mapping = {k.lower(): (v.lower() if isinstance(v, str) else v)
                                    for k, v in span_type_mapping.items()}
        except Exception:
            pass

    if nodes is None:
        nodes = load_nodes(excel_path)
    if conns is None:
        conns = load_connections_with_attrs(excel_path)
    if sections is None:
        sections = load_sections(excel_path)
    if not conns or not nodes:
        return []
    if not sections:
        _log("  Spans comparison: no sections sheet or empty")
        return []

    # connection_id -> raw count per canonical span type (after mapping POA values)
    conn_id_to_poa_counts: Dict[str, Dict[str, int]] = {}
    for sec in sections:
        cid = sec.get("connection_id")
        if not cid:
            continue
        poas = sec.get("POAs") or []
        if cid not in conn_id_to_poa_counts:
            conn_id_to_poa_counts[cid] = {}
        for t in poas:
            poa_name = (t or "").strip().lower()
            if not poa_name:
                continue
            canonical = span_type_mapping.get(poa_name, poa_name)
            if canonical is None:
                continue
            if not isinstance(canonical, str):
                canonical = str(canonical).lower()
            key = canonical.strip().lower()
            conn_id_to_poa_counts[cid][key] = conn_id_to_poa_counts[cid].get(key, 0) + 1

    scid_to_pplx = _build_scid_to_pplx_map(pplx_file_paths, extract_scid_fn)

    from src.core.handler import PPLXHandler
    if pplx_cache is None:
        pplx_cache = {}
    # One tree walk per path: path -> [(span_type, length_in), ...]
    path_span_pairs: Dict[str, List[Tuple[str, float]]] = {}
    rows: List[Dict[str, Any]] = []
    t_spans_loop = time.perf_counter()
    t_pplx_load_total = 0.0
    pplx_load_count = 0

    # Optional: midspan heights workbook for comm attachment counts
    midspan_counts: Dict[Tuple[str, str], int] = {}
    comm_types = {"catv", "fiber", "telco"}
    if midspan_heights_path is not None:
        try:
            midspan_counts = load_midspan_heights_counts(midspan_heights_path, log_callback=_log)
        except Exception as e:
            _log(f"  Midspan heights: error loading {midspan_heights_path}: {e}")
            midspan_counts = {}

    # Track which span indices have been consumed per PPLX file so each
    # PPLX span is matched to at most one Katapult connection.
    consumed_span_indices: Dict[str, Set[int]] = {}

    def _load_pplx_pairs(pplx_path: str) -> Optional[List[Tuple[str, float, Optional[float]]]]:
        """Load handler + span triples (type, length, angle) for a PPLX path, caching results."""
        nonlocal t_pplx_load_total, pplx_load_count
        if pplx_path not in pplx_cache:
            try:
                t_load = time.perf_counter()
                pplx_cache[pplx_path] = PPLXHandler(pplx_path)
                t_pplx_load_total += time.perf_counter() - t_load
                pplx_load_count += 1
            except Exception:
                pplx_cache[pplx_path] = None
        handler = pplx_cache[pplx_path]
        if handler is None or handler.root is None:
            return None
        if pplx_path not in path_span_pairs:
            path_span_pairs[pplx_path] = handler.get_span_type_length_angle_triples_for_spans_qc()
        return path_span_pairs[pplx_path]

    def _count_and_consume(pplx_path: str, triples: List[Tuple[str, float, Optional[float]]],
                           length_in: float, tol_in: float, target_angle: Optional[float] = None
                           ) -> Tuple[Dict[str, int], Set[str]]:
        """
        Count matching unconsumed spans per type and consume them.
        For communication spans (CATV, Fiber, Telco), match using both length and angle.
        For power spans (Primary, Neutral, Secondary), match using length only.
        Also returns a set of span types where angle matches (within 10 deg) but length does not —
        these are candidates for 'Length' QC status instead of 'FAIL'.
        """
        if pplx_path not in consumed_span_indices:
            consumed_span_indices[pplx_path] = set()
        consumed = consumed_span_indices[pplx_path]
        counts: Dict[str, int] = {}
        angle_only_types: Set[str] = set()
        matched: List[int] = []
        angle_tolerance_rad = math.radians(10)  # 10 degrees

        for idx, (stype, L, angle) in enumerate(triples):
            if idx in consumed:
                continue

            stype_key = stype.lower()
            length_ok = abs(L - length_in) <= tol_in

            # Angle check (all span types when target_angle and angle are available)
            angle_ok = True
            if target_angle is not None and angle is not None:
                angle_diff = abs(angle - target_angle)
                if angle_diff > math.pi:
                    angle_diff = 2 * math.pi - angle_diff
                angle_ok = angle_diff <= angle_tolerance_rad

            # When both target_angle and span angle are available, use angle+length for all types.
            # Without angle info, fall back to length-only for power spans.
            use_angle = target_angle is not None and angle is not None
            if use_angle:
                # Angle+length matching (comm and power)
                if angle_ok and length_ok:
                    counts[stype_key] = counts.get(stype_key, 0) + 1
                    matched.append(idx)
                elif angle_ok and not length_ok:
                    angle_only_types.add(stype_key)
            elif stype_key not in comm_types:
                # Power spans with no angle info: match by length only
                if length_ok:
                    counts[stype_key] = counts.get(stype_key, 0) + 1
                    matched.append(idx)

        consumed.update(matched)
        return counts, angle_only_types

    for c in conns:
        conn_id = c.get("connection_id")
        if not conn_id or conn_id not in conn_id_to_poa_counts:
            continue
        span_dist = c.get("span_distance")
        if span_dist is None or span_dist <= 0:
            continue
        nid1, nid2 = c["node_id_1"], c["node_id_2"]
        n1 = nodes.get(nid1)
        n2 = nodes.get(nid2)
        if not n1 or not n2:
            continue
        scid1 = str(n1.get("scid") or "").strip()
        scid2 = str(n2.get("scid") or "").strip()
        base1 = (scid1.split() or [""])[0]
        base2 = (scid2.split() or [""])[0]
        if not base1 or not base2:
            continue
        s1, s2 = scid1 or base1, scid2 or base2
        if _pole_order(s1) <= _pole_order(s2):
            pole, to_pole = s1, s2
        else:
            pole, to_pole = s2, s1

        katapult_counts = dict(conn_id_to_poa_counts[conn_id])

        # Adjust comm (CATV/Fiber/Telco) counts using midspan heights when available.
        raw_comm = {t: katapult_counts.get(t, 0) for t in comm_types if katapult_counts.get(t, 0) > 0}
        if raw_comm:
            key_pair = tuple(sorted([base1, base2]))
            limit = midspan_counts.get(key_pair)
            if limit is None:
                adj_comm = dict(raw_comm)
            else:
                total_raw = sum(raw_comm.values())
                target = min(limit, total_raw)
                if target <= 0:
                    adj_comm = {t: 0 for t in comm_types}
                elif target == total_raw:
                    adj_comm = raw_comm
                else:
                    types_sorted = sorted(raw_comm.keys(), key=lambda t: raw_comm[t], reverse=True)
                    adj_comm = {t: 0 for t in comm_types}
                    remaining = target
                    for t in types_sorted:
                        if remaining <= 0:
                            break
                        if raw_comm[t] > 0:
                            adj_comm[t] = min(raw_comm[t], 1)
                            remaining -= adj_comm[t]
                    idx = 0
                    while remaining > 0 and types_sorted:
                        t = types_sorted[idx % len(types_sorted)]
                        spare = raw_comm[t] - adj_comm[t]
                        if spare > 0:
                            adj_comm[t] += 1
                            remaining -= 1
                        idx += 1
            for t in comm_types:
                new_val = adj_comm.get(t, 0)
                if new_val <= 0:
                    katapult_counts.pop(t, None)
                else:
                    katapult_counts[t] = new_val

        pole_base = _base_pole_id(pole)
        to_pole_base = _base_pole_id(to_pole)
        length_in = float(span_dist) * 12.0
        tol_in = float(length_tolerance_in)

        # Node coordinates (n1/n2 from Katapult connection order, may differ from pole/to_pole order)
        lat1 = n1.get("latitude"); lon1 = n1.get("longitude")
        lat2 = n2.get("latitude"); lon2 = n2.get("longitude")
        # Identify which node belongs to `pole` (the lower-ordered one after reordering)
        pole_node, other_node = (n1, n2) if pole == s1 else (n2, n1)
        pole_lat = pole_node.get("latitude"); pole_lon = pole_node.get("longitude")
        other_lat = other_node.get("latitude"); other_lon = other_node.get("longitude")

        # Use ONE PPLX file per connection (primary pole first, fallback to secondary).
        # Bearing must be from the PPLX pole outward toward the other pole,
        # because PPLX spans point away from their home pole.
        pplx_counts: Dict[str, int] = {}
        pplx_path = scid_to_pplx.get(pole_base)
        if pplx_path:
            # PPLX file belongs to `pole`; bearing = pole -> to_pole
            src_lat, src_lon, dst_lat, dst_lon = pole_lat, pole_lon, other_lat, other_lon
        else:
            pplx_path = scid_to_pplx.get(to_pole_base)
            if pplx_path:
                # Fallback: PPLX file belongs to `to_pole`; bearing = to_pole -> pole
                src_lat, src_lon, dst_lat, dst_lon = other_lat, other_lon, pole_lat, pole_lon
        target_angle: Optional[float] = None
        if pplx_path and src_lat is not None and src_lon is not None and dst_lat is not None and dst_lon is not None:
            try:
                target_angle = bearing_rad_from_lat_lon(src_lat, src_lon, dst_lat, dst_lon)
            except Exception:
                pass
        angle_only_types: Set[str] = set()
        if pplx_path:
            triples = _load_pplx_pairs(pplx_path)
            if triples is not None:
                pplx_counts, angle_only_types = _count_and_consume(pplx_path, triples, length_in, tol_in, target_angle)

        all_types = set(katapult_counts) | set(pplx_counts)
        for span_type in sorted(all_types):
            k = katapult_counts.get(span_type, 0)
            p = pplx_counts.get(span_type, 0)
            if k == p:
                qc = "PASS"
            elif k > p and span_type in angle_only_types:
                qc = "Length"
            else:
                qc = "FAIL"
            rows.append({
                "Pole": pole,
                "To Pole": to_pole,
                "Span Type": span_type.title(),
                "Katapult": k,
                "PPLX": p,
                "QC": qc,
            })

    rows.sort(key=lambda r: (leading_int(r["Pole"]), leading_int(r["To Pole"]), r["Span Type"]))
    t_spans_loop_elapsed = time.perf_counter() - t_spans_loop
    _log(f"  Spans comparison: {len(rows)} rows in {t_spans_loop_elapsed:.1f}s (tol: {length_tolerance_in:.0f} in, PPLX loads: {pplx_load_count}/{len(pplx_cache)}, files with span data: {len(path_span_pairs)})")
    return rows


def main() -> None:
    base = Path(__file__).parent
    excel_path = base / "Test Files" / "OPPD" / "NEOM104 Nodes-Sections-Connections XLSX.xlsx"
    prj_path = base / "data" / "OPPD" / "shape" / "ElectricLine selection.prj"
    shp_path = base / "data" / "OPPD" / "shape" / "ElectricLine selection.shp"

    if not excel_path.exists():
        print(f"Excel not found: {excel_path}")
        return
    if not shp_path.exists() or not prj_path.exists():
        print(f"Shapefile/PRJ not found: {shp_path} / {prj_path}")
        return

    nodes = load_nodes(excel_path)
    connections = load_connections(excel_path)
    transformer = transformer_wgs84_to_layer(prj_path)

    print(f"Nodes with coords: {len(nodes)}, Connections: {len(connections)}")
    print("Pole Number (SCID) 1->2, 2->3, etc.: wire spec between those poles\n")

    results: List[Tuple[str, str, Optional[str], Optional[str], float, float, Dict[str, Any]]] = []
    skipped = 0
    for nid1, nid2 in connections:
        n1 = nodes.get(nid1)
        n2 = nodes.get(nid2)
        if not n1 or not n2:
            skipped += 1
            continue
        lat1 = n1.get("latitude")
        lon1 = n1.get("longitude")
        lat2 = n2.get("latitude")
        lon2 = n2.get("longitude")
        if lat1 is None or lon1 is None or lat2 is None or lon2 is None:
            skipped += 1
            continue
        try:
            lat1, lon1 = float(lat1), float(lon1)
            lat2, lon2 = float(lat2), float(lon2)
        except (TypeError, ValueError):
            skipped += 1
            continue
        scid1 = n1.get("scid")
        scid2 = n2.get("scid")
        if scid1 is not None:
            scid1 = str(scid1).strip()
        if scid2 is not None:
            scid2 = str(scid2).strip()

        spec = wire_spec_between_points(lat1, lon1, lat2, lon2, prj_path, shp_path, transformer)
        results.append((scid1 or nid1, scid2 or nid2, nid1, nid2, lat1, lon1, spec))

    # Sort by (scid1, scid2) when numeric so we get 1->2, 2->3 order for display
    def sort_key(r):
        s1, s2 = r[0], r[1]
        try:
            a, b = int(s1), int(s2)
            return (0, a, b)
        except (ValueError, TypeError):
            return (1, str(s1), str(s2))

    results.sort(key=sort_key)

    print(f"Skipped (no coords): {skipped}\n")
    print(f"{'Pole (SCID)':<12} {'->':^4} {'Pole (SCID)':<12} {'d_masterma (primary)':<24} {'d_neutralm':<14} {'d_orientat':<8} {'dist1_ft':<10} {'dist2_ft':<10}")
    print("-" * 100)
    for scid1, scid2, _n1, _n2, _la1, _lo1, spec in results:
        master = (spec.get("d_masterma") or "").strip() or "-"
        neutral = (spec.get("d_neutralm") or "").strip() or "-"
        orient = (spec.get("d_orientat") or "").strip() or "-"
        d1 = spec.get("dist1_ft")
        d2 = spec.get("dist2_ft")
        d1s = f"{d1:.1f}" if d1 is not None else "-"
        d2s = f"{d2:.1f}" if d2 is not None else "-"
        print(f"{str(scid1):<12} {'->':^4} {str(scid2):<12} {master:<24} {neutral:<14} {orient:<8} {d1s:<10} {d2s:<10}")

    print("\nDone.")


if __name__ == "__main__":
    main()
