#!/usr/bin/env python3
"""
Simulate app run to generate change_log.xlsx (Aux Data + Wire Specs).
Uses: Excel from config/state, PPLX from Test Files/OPPD/PPLX, OPPD shapefiles.
Does not modify PPLX files; writes change_log to project root.
"""
import os
import sys
from pathlib import Path

# Project root on path
ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from src.core.handler import PPLXHandler
from src.core.logic import extract_scid_from_filename
from src.excel.loader import load_excel_data


def main():
    excel_path = ROOT / "Test Files" / "OPPD" / "NEOM104 Nodes-Sections-Connections XLSX.xlsx"
    pplx_dir = ROOT / "Test Files" / "OPPD" / "PPLX"
    shape_base = ROOT / "data" / "OPPD" / "shape"
    out_path = ROOT / "change_log.xlsx"

    if not excel_path.exists():
        print(f"Excel not found: {excel_path}")
        return 1
    if not pplx_dir.is_dir():
        print(f"PPLX dir not found: {pplx_dir}")
        return 1

    excel_data = load_excel_data(str(excel_path), log_callback=print)
    valid_scids = set(excel_data.keys()) if excel_data else set()
    print(f"Excel: {len(valid_scids)} poles")

    files = sorted(pplx_dir.glob("*.pplx"))
    if not files:
        print("No PPLX files found")
        return 1
    file_paths = [str(f) for f in files]
    print(f"PPLX files: {len(file_paths)}")

    # --- Aux Data sheet (read from PPLX + Excel) ---
    csv_data = []
    for i, path in enumerate(file_paths):
        filename = os.path.basename(path)
        scid = extract_scid_from_filename(filename)
        if excel_data and scid not in valid_scids:
            continue
        try:
            handler = PPLXHandler(path)
            aux = handler.get_aux_data()
            mr_note = excel_data.get(scid, {}).get("mr_note", "") if excel_data and scid in excel_data else ""
            csv_data.append({
                "File Name": filename,
                "MR Note": mr_note,
                "Aux Data 1": aux.get("Aux Data 1", "Unset"),
                "Aux Data 2": aux.get("Aux Data 2", "Unset"),
                "Aux Data 3": aux.get("Aux Data 3", "Unset"),
                "Aux Data 4": aux.get("Aux Data 4", "Unset"),
                "Aux Data 5": aux.get("Aux Data 5", "Unset"),
            })
        except Exception as e:
            print(f"  Skip {filename}: {e}")
    print(f"Aux Data rows: {len(csv_data)}")

    # --- Wire Specs sheet (OPPD: Pole-Pole, Wire_Type, PPLX, Shape) ---
    wire_spec_data = []
    if shape_base.exists():
        try:
            from wire_spec_from_excel import build_wire_spec_comparison
            # Limit connections for quick run (set to None for full run)
            wire_spec_data = build_wire_spec_comparison(
                excel_path, file_paths, shape_base, extract_scid_from_filename,
                max_connections=40,
            )
            print(f"Wire Specs rows: {len(wire_spec_data)}")
        except Exception as e:
            print(f"Wire spec comparison failed: {e}")
    else:
        print("Shape path not found, skipping Wire Specs")

    # --- Write change_log.xlsx ---
    wb = Workbook()
    wb.remove(wb.active)

    ws_aux = wb.create_sheet("Aux Data", 0)
    aux_headers = ["File Name", "MR Note", "Aux Data 1", "Aux Data 2", "Aux Data 3", "Aux Data 4", "Aux Data 5"]
    for col, h in enumerate(aux_headers, 1):
        ws_aux.cell(row=1, column=col, value=h)
    for row_idx, row_dict in enumerate(csv_data, start=2):
        for col, h in enumerate(aux_headers, 1):
            ws_aux.cell(row=row_idx, column=col, value=row_dict.get(h, ""))

    ws_wire = wb.create_sheet("Wire Specs", 1)
    wire_headers = ["Pole-Pole", "Wire_Type", "PPLX", "Shape"]
    for col, h in enumerate(wire_headers, 1):
        ws_wire.cell(row=1, column=col, value=h)
    for row_idx, row_dict in enumerate(wire_spec_data, start=2):
        for col, h in enumerate(wire_headers, 1):
            ws_wire.cell(row=row_idx, column=col, value=row_dict.get(h, ""))

    wb.save(out_path)
    print(f"Saved: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
